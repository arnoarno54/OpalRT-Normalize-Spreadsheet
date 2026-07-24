"""
Opal RT Spreadsheet Cleaner
---------------------------
Internal tool for converting messy lead spreadsheets into a Dynamics-compatible
import CSV.

Built by Arnaud Joakim <arnaud.joakim@opal-rt.com>
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# PAGE CONFIG - must be the first Streamlit call
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Opal RT Spreadsheet Cleaner",
    page_icon="🔷",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ===========================================================================
# CONSTANTS
# ===========================================================================

# Exact Dynamics export column order (do not change)
DYNAMICS_COLUMNS: List[str] = [
    "(Do Not Modify) Lead",
    "(Do Not Modify) Row Checksum",
    "(Do Not Modify) Modified On",
    "Subject",
    "First Name",
    "Last Name",
    "Job Title",
    "Company Name",
    "Email",
    "Business Phone",
    "Country",
    "State or Province",
    "Description",
    "Lead Source",
    "Rating",
    "Source Campaign",
    "Market Segment",
    "Main Application",
    "Industry Sector",
    "LinkedIn",
    "Allow Marketing Communication",
]

MANDATORY_FIELDS: List[str] = [
    "Subject",
    "First Name",
    "Last Name",
    "Email",
    "Company Name",
    "Country",
]

FIELD_MAX_LENGTHS: Dict[str, int] = {
    "First Name": 58,
    "Last Name": 50,
    "Company Name": 100,
    "Job Title": 100,
    "Email": 100,
    "LinkedIn": 500,
    "Description": 2000,
    "Subject": 300,
    "Business Phone": 50,
}

# Dropdown options (exact values from ImportLeadTemplate.xlsm)
LEAD_SOURCE_OPTIONS = [
    "",
    "Shows",
    "Web",
    "Prospection",
    "Webinar",
    "Referral",
    "Social Media",
    "Customer Portal",
    "SPS",
    "Others",
]

RATING_OPTIONS = ["", "Cold", "Warm", "Hot"]

ALLOW_MARKETING_OPTIONS = ["", "Yes", "No"]

INDUSTRY_SECTOR_OPTIONS = [
    "",
    "Academic - Research or Post-graduate",
    "Academic - Undergraduate",
    "Consulting & Engineering Firm",
    "Defense",
    "Electrical Utility",
    "Manufacturer",
    "Other",
    "Research Lab - Industrial & Gov.",
    "Stock - Inventory",
]

MARKET_SEGMENT_OPTIONS = [
    "",
    "Aerospace",
    "Automotive",
    "Energy Conversion",
    "Marine, Railway, Off-Highway",
    "Power System",
]

MAIN_APPLICATION_BY_SEGMENT: Dict[str, List[str]] = {
    "": [""],
    "Aerospace": [
        "",
        "Autonomous Systems (Aero)",
        "Avionics System",
        "Electrical Actuators and Servos",
        "EVTOL",
        "More Electrical Aircraft",
        "Onboard System",
        "Other (if nothing fits) Aero",
        "Propulsion and APU",
        "Testbench - Test Automation and Monitoring from RTS",
    ],
    "Automotive": [
        "",
        "Autonomous Systems (Auto)",
        "Body & Chassis",
        "Charging",
        "EV/HEV Powertrain",
        "Full Vehicle Simulation",
        "ICE Powertrain",
        "Other (if nothing fits) Auto",
    ],
    "Energy Conversion": [
        "",
        "Autonomous Systems (Energy Conversion)",
        "Backup Power (UPS)",
        "Inverter/Converter",
        "Medium and Large Drive (>150KW)",
        "Other (if nothing fits) EnergyConversion",
        "Small Drive (<150KW)",
    ],
    "Marine, Railway, Off-Highway": [
        "",
        "Autonomous Systems (Marine, Railway, Off-Highway)",
        "BMS Control",
        "Grid Infrastructure",
        "Onboard Power System",
        "Other (if nothing fits) Marine, Railway, Off-Highway",
        "Propulsion Control",
    ],
    "Power System": [
        "",
        "Autonomous Systems (Power Systems)",
        "Conventional Generation",
        "Converter-Based Energy Resource",
        "Distribution",
        "FACTS & HVDC",
        "Microgrid",
        "Other (if nothing fits) PowerSystem",
        "Substation",
        "Transmission",
    ],
}

# Canonical Dynamics countries (244 - exact list from template)
COUNTRIES: List[str] = [
    "Afghanistan", "African Country (non-maghrebian)", "Aland Island", "Albania",
    "Algeria", "American Samoa", "Andorra", "Angola", "Anguilla", "Antartica",
    "Antigua and Barbuda", "Argentina", "Armenia", "Aruba", "Australia", "Austria",
    "Azerbaijan", "Bahamas", "Bahrain", "Bangladesh", "Barbados", "Belarus",
    "Belgium", "Belize", "Benin", "Bermuda", "Bhutan", "Bolivia",
    "Bosnia and Herzegovina", "Botswana", "Bouvet Island", "Brazil",
    "British Indian Ocean Territory", "Brunei Darussalam", "Bulgaria",
    "Burkina Faso", "Burundi", "Cambodia", "Cameroon", "Canada", "Cape Verde",
    "Cayman Islands", "Central African Republic", "Chad", "Chile", "China",
    "Chrismas Island", "Cocos (Keeling) Islands", "Colombia", "Comoros", "Congo",
    "Congo, The democatic Republic of the", "Cook Islands", "Costa Rica",
    "Croatia", "Cuba", "Cyprus", "Czech Republic", "Denmark", "Djibouti",
    "Dominica", "Dominican Republic", "Egypt", "El Salvador", "Ecuador",
    "Equatorial Guinea", "Eritrea", "Estonia", "Ethiopia",
    "Falkland Islands (Malvinas)", "Faroe Island", "Fiji", "Finland", "France",
    "French Guiana", "French Polynesia", "French Southern Territories",
    "French-Guadeloupe", "French-Martinique", "Gabon", "Gambia", "Georgia",
    "Germany", "Ghana", "Gibraltar", "Greece", "Greenland", "Grenada", "Guam",
    "Guatemala", "Guernser", "Guinea", "Guinea-Bissau", "Guyana", "Haiti",
    "Heard Island and McDonals Islands", "Honduras", "Hong Kong", "Hungary",
    "Iceland", "India", "Indonesia", "Iran (Islamic Republic of)", "Iraq",
    "Ireland", "Isle of Man", "Israel", "Italy", "Ivory Coast", "Jamaica",
    "Japan", "Jordan", "Kazakhstan", "Kenya", "Kiribati", "Kuwait", "Kyrgyzstan",
    "Lakshadweep", "Lao People's Democratic republic", "Latvia", "Lebanon",
    "Lesotho", "Liberia", "Libya", "Liechtenstein", "Lithuania", "Luxembourg",
    "Macao", "Macedonia", "Madagascar", "Malawi", "Malaysia", "Maldives", "Mali",
    "Malta", "Marshall Islands", "Mauritania", "Mauritius", "Mayotte", "Mexico",
    "Micronesia, Federated States of", "Moldova", "Monaco", "Mongolia",
    "Montenegro", "Montserrat", "Morocco", "Mozambique", "Myanmar", "N/A",
    "Namibia", "Nepal", "Netherlands", "Netherlands Antilles", "New Caledonia",
    "New Zealand", "Nicaragua", "Niger", "Nigeria", "Niue", "Norfolk Iskand",
    "Northern Mariana Islands", "Norway", "Oman", "Pakistan", "Palau",
    "Palestine", "Panama", "Papua New Guinea", "Paraguay", "Peru", "Philippines",
    "Pitcairn", "Poland", "Portugal", "Puerto Rico", "Qatar", "Reunion",
    "Romania", "Russia", "Rwanda", "Saint Barthelemy", "Saint Helena",
    "Saint Kitts and Nevis", "Saint Lucia", "Saint Pierre and Miquelon",
    "Saint Vincent and the Grenadines", "Samoa", "San Marino",
    "Sao Tome and Principe", "Saudi Arabia", "Senegal", "Serbia", "Seychelles",
    "Shanghai", "Sierra Leone", "Singapore", "Slovakia", "Slovenia",
    "Solomon Islands", "Somalia", "South Africa",
    "South Georgia and the South Sandwich Islands", "South Korea", "Spain",
    "Sri Lanka", "St Martin", "Sudan", "Suriname", "Svalbard and Jan Mayen",
    "Swaziland", "Sweden", "Switzerland", "Syria", "Taiwan", "Tajikistan",
    "Tanzania", "Thailand", "Timor-Leste", "Togo", "Trinidad and Tobago",
    "Tunisia", "Turkey", "Turkmenistan", "Turks and Caicos Island", "Tuvalu",
    "Uganda", "Ukraine", "United Arab Emirates", "United Kingdom",
    "United States", "Uruguay", "Uzbekistan", "Vanuatu", "Vatican City State",
    "Venezuela", "Vietnam", "Virgin Islands, British", "Virgin Islands, U.S.",
    "Wallis and Futuna", "Western Sahara", "Yemen", "Zambia", "Zimbabwe",
]

# Common country aliases → canonical name (lowercase keys)
COUNTRY_ALIASES: Dict[str, str] = {
    "usa": "United States",
    "u.s.a.": "United States",
    "u.s.a": "United States",
    "us": "United States",
    "u.s.": "United States",
    "u.s": "United States",
    "united states of america": "United States",
    "united states": "United States",
    "america": "United States",
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "u.k": "United Kingdom",
    "great britain": "United Kingdom",
    "britain": "United Kingdom",
    "england": "United Kingdom",
    "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "northern ireland": "United Kingdom",
    "uae": "United Arab Emirates",
    "u.a.e.": "United Arab Emirates",
    "u.a.e": "United Arab Emirates",
    "russian federation": "Russia",
    "russia": "Russia",
    "south korea": "South Korea",
    "korea, republic of": "South Korea",
    "republic of korea": "South Korea",
    "korea (south)": "South Korea",
    "rok": "South Korea",
    "viet nam": "Vietnam",
    "vietnam": "Vietnam",
    "ivory coast": "Ivory Coast",
    "côte d'ivoire": "Ivory Coast",
    "cote d'ivoire": "Ivory Coast",
    "czech republic": "Czech Republic",
    "czechia": "Czech Republic",
    "hong kong sar": "Hong Kong",
    "hk": "Hong Kong",
    "taiwan, province of china": "Taiwan",
    "republic of china": "Taiwan",
    "iran": "Iran (Islamic Republic of)",
    "islamic republic of iran": "Iran (Islamic Republic of)",
    "laos": "Lao People's Democratic republic",
    "macau": "Macao",
    "syrian arab republic": "Syria",
    "republic of moldova": "Moldova",
    "bolivia, plurinational state of": "Bolivia",
    "venezuela, bolivarian republic of": "Venezuela",
    "tanzania, united republic of": "Tanzania",
    "north macedonia": "Macedonia",
    "republic of north macedonia": "Macedonia",
    "swaziland": "Swaziland",
    "eswatini": "Swaziland",
    "myanmar (burma)": "Myanmar",
    "burma": "Myanmar",
    "the netherlands": "Netherlands",
    "holland": "Netherlands",
    "deutschland": "Germany",
    "espana": "Spain",
    "españa": "Spain",
    "italia": "Italy",
    "brasil": "Brazil",
    "republic of singapore": "Singapore",
    "kingdom of saudi arabia": "Saudi Arabia",
    "ksa": "Saudi Arabia",
    "people's republic of china": "China",
    "prc": "China",
    "mainland china": "China",
}

# US states (with abbreviations)
US_STATES_FULL: List[str] = [
    "Alabama", "Alaska", "American Samoa", "Arizona", "Arkansas", "California",
    "Colorado", "Connecticut", "Delaware", "District of Columbia", "Florida",
    "Georgia", "Guam", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa",
    "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts",
    "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska",
    "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Northern Mariana Islands", "Ohio",
    "Oklahoma", "Oregon", "Pennsylvania", "Puerto Rico", "Rhode Island",
    "South Carolina", "South Dakota", "Tennessee", "Texas",
    "United States Minor Outlying Islands", "Utah", "Vermont",
    "Virgin Islands, U.S.", "Virginia", "Washington", "West Virginia",
    "Wisconsin", "Wyoming",
]

US_STATE_ABBR: Dict[str, str] = {
    "al": "Alabama", "ak": "Alaska", "as": "American Samoa", "az": "Arizona",
    "ar": "Arkansas", "ca": "California", "co": "Colorado", "ct": "Connecticut",
    "de": "Delaware", "dc": "District of Columbia", "fl": "Florida",
    "ga": "Georgia", "gu": "Guam", "hi": "Hawaii", "id": "Idaho",
    "il": "Illinois", "in": "Indiana", "ia": "Iowa", "ks": "Kansas",
    "ky": "Kentucky", "la": "Louisiana", "me": "Maine", "md": "Maryland",
    "ma": "Massachusetts", "mi": "Michigan", "mn": "Minnesota",
    "ms": "Mississippi", "mo": "Missouri", "mt": "Montana", "ne": "Nebraska",
    "nv": "Nevada", "nh": "New Hampshire", "nj": "New Jersey",
    "nm": "New Mexico", "ny": "New York", "nc": "North Carolina",
    "nd": "North Dakota", "mp": "Northern Mariana Islands", "oh": "Ohio",
    "ok": "Oklahoma", "or": "Oregon", "pa": "Pennsylvania", "pr": "Puerto Rico",
    "ri": "Rhode Island", "sc": "South Carolina", "sd": "South Dakota",
    "tn": "Tennessee", "tx": "Texas", "ut": "Utah", "vt": "Vermont",
    "vi": "Virgin Islands, U.S.", "va": "Virginia", "wa": "Washington",
    "wv": "West Virginia", "wi": "Wisconsin", "wy": "Wyoming",
}

# Canadian provinces (with abbreviations) - using template's spelling (Québec)
CA_PROVINCES_FULL: List[str] = [
    "Alberta", "British Columbia", "Manitoba", "New Brunswick",
    "Newfoundland and Labrador", "Northwest Territories", "Nova Scotia",
    "Nunavut", "Ontario", "Prince Edward Island", "Québec", "Saskatchewan",
    "Yukon Territory",
]

CA_PROVINCE_ABBR: Dict[str, str] = {
    "ab": "Alberta", "bc": "British Columbia", "mb": "Manitoba",
    "nb": "New Brunswick", "nl": "Newfoundland and Labrador",
    "nt": "Northwest Territories", "ns": "Nova Scotia", "nu": "Nunavut",
    "on": "Ontario", "pe": "Prince Edward Island", "pei": "Prince Edward Island",
    "qc": "Québec", "que": "Québec", "quebec": "Québec",
    "sk": "Saskatchewan", "yt": "Yukon Territory", "yk": "Yukon Territory",
}

# Source-column detection rules. Order matters for tie-breaking (specific first).
COLUMN_ALIASES: Dict[str, List[str]] = {
    "First Name": [
        "first name", "firstname", "first_name", "fname", "given name",
        "given_name", "givenname", "first", "prenom", "prénom", "nom prenom",
    ],
    "Last Name": [
        "last name", "lastname", "last_name", "lname", "surname",
        "family name", "family_name", "familyname", "last", "nom de famille",
    ],
    "Company Name": [
        "company name", "company", "companyname", "company_name", "organization",
        "organisation", "org", "employer", "business name", "account", "firm",
        "entreprise", "société", "societe",
    ],
    "Job Title": [
        "job title", "jobtitle", "job_title", "title", "position", "role",
        "job role", "designation", "fonction", "poste",
    ],
    "Email": [
        "email", "e-mail", "email address", "emailaddress", "email_address",
        "work email", "business email", "corporate email", "professional email",
        "primary email", "mail", "courriel", "adresse email",
    ],
    "Business Phone": [
        "business phone", "businessphone", "business_phone", "work phone",
        "workphone", "work_phone", "office phone", "company phone",
        "mobile phone", "mobilephone", "mobile_phone", "mobile", "cell",
        "cell phone", "cellphone", "phone", "telephone", "tel", "phone number",
        "contact number", "phone no", "primary phone", "téléphone",
    ],
    "LinkedIn": [
        "linkedin", "linkedin profile", "linkedin profile url",
        "linkedin url", "linkedin link", "linkedinprofile", "linkedin_profile",
        "linkedin profile name", "li profile", "li url",
    ],
    "Country": [
        "country", "country/region", "country region", "country_region",
        "nation", "pays",
    ],
    "State or Province": [
        "state or province", "state/province", "state province", "state",
        "province", "region", "state_province", "stateprovince",
    ],
    "Description": [
        "description", "notes", "note", "comments", "comment", "remarks",
        "details", "about",
    ],
    "Market Segment": [
        "market segment", "marketsegment", "segment", "market",
    ],
    "Main Application": [
        "main application", "mainapplication", "application", "use case",
        "usecase",
    ],
    "Industry Sector": [
        "industry sector", "industrysector", "industry", "sector", "vertical",
    ],
}

# "Location" gets a separate slot because it's parsed differently
LOCATION_ALIASES: List[str] = [
    "location", "city", "addresse", "address", "city/state", "city, state",
    "geography", "geo", "city and country",
]

# ---------------------------------------------------------------------------
# City → (Country, State/Province) lookup
# ---------------------------------------------------------------------------
# Used as a final fallback in parse_location_string for LinkedIn-style strings
# like 'Greater Chicago Area', 'Greater Toulouse Metropolitan Area', or just
# 'Houston'. Keys are LOWERCASE and ASCII-folded (no accents).
#
# Rules of curation:
#   - State/Province only included for US and Canada (matches template).
#   - Ambiguous city names (Springfield, Cambridge, London-the-CA-city, etc.)
#     are intentionally OMITTED rather than guessed.
#   - For mildly ambiguous names (Birmingham, Manchester, Athens), the most
#     populous / globally-recognised version wins.
CITY_TO_GEO: Dict[str, Tuple[str, str]] = {
    # ---------- United States ----------
    "new york": ("United States", "New York"),
    "new york city": ("United States", "New York"),
    "nyc": ("United States", "New York"),
    "los angeles": ("United States", "California"),
    "chicago": ("United States", "Illinois"),
    "houston": ("United States", "Texas"),
    "phoenix": ("United States", "Arizona"),
    "philadelphia": ("United States", "Pennsylvania"),
    "san antonio": ("United States", "Texas"),
    "san diego": ("United States", "California"),
    "dallas": ("United States", "Texas"),
    "fort worth": ("United States", "Texas"),
    "san jose": ("United States", "California"),
    "austin": ("United States", "Texas"),
    "jacksonville": ("United States", "Florida"),
    "columbus": ("United States", "Ohio"),
    "charlotte": ("United States", "North Carolina"),
    "san francisco": ("United States", "California"),
    "silicon valley": ("United States", "California"),
    "indianapolis": ("United States", "Indiana"),
    "seattle": ("United States", "Washington"),
    "denver": ("United States", "Colorado"),
    "washington dc": ("United States", "District of Columbia"),
    "washington d.c.": ("United States", "District of Columbia"),
    "boston": ("United States", "Massachusetts"),
    "nashville": ("United States", "Tennessee"),
    "el paso": ("United States", "Texas"),
    "detroit": ("United States", "Michigan"),
    "memphis": ("United States", "Tennessee"),
    "oklahoma city": ("United States", "Oklahoma"),
    "las vegas": ("United States", "Nevada"),
    "louisville": ("United States", "Kentucky"),
    "baltimore": ("United States", "Maryland"),
    "milwaukee": ("United States", "Wisconsin"),
    "albuquerque": ("United States", "New Mexico"),
    "tucson": ("United States", "Arizona"),
    "fresno": ("United States", "California"),
    "sacramento": ("United States", "California"),
    "atlanta": ("United States", "Georgia"),
    "kansas city": ("United States", "Missouri"),
    "colorado springs": ("United States", "Colorado"),
    "miami": ("United States", "Florida"),
    "fort lauderdale": ("United States", "Florida"),
    "raleigh": ("United States", "North Carolina"),
    "omaha": ("United States", "Nebraska"),
    "long beach": ("United States", "California"),
    "virginia beach": ("United States", "Virginia"),
    "oakland": ("United States", "California"),
    "minneapolis": ("United States", "Minnesota"),
    "saint paul": ("United States", "Minnesota"),
    "st. paul": ("United States", "Minnesota"),
    "tulsa": ("United States", "Oklahoma"),
    "tampa": ("United States", "Florida"),
    "new orleans": ("United States", "Louisiana"),
    "cleveland": ("United States", "Ohio"),
    "pittsburgh": ("United States", "Pennsylvania"),
    "cincinnati": ("United States", "Ohio"),
    "saint louis": ("United States", "Missouri"),
    "st. louis": ("United States", "Missouri"),
    "orlando": ("United States", "Florida"),
    "salt lake city": ("United States", "Utah"),
    "buffalo": ("United States", "New York"),
    "anaheim": ("United States", "California"),
    "santa ana": ("United States", "California"),
    "irvine": ("United States", "California"),
    "berkeley": ("United States", "California"),
    "palo alto": ("United States", "California"),
    "mountain view": ("United States", "California"),
    "sunnyvale": ("United States", "California"),
    "santa clara": ("United States", "California"),
    "santa monica": ("United States", "California"),
    "cupertino": ("United States", "California"),
    "menlo park": ("United States", "California"),
    "redwood city": ("United States", "California"),
    "san mateo": ("United States", "California"),
    "santa barbara": ("United States", "California"),
    "burbank": ("United States", "California"),
    "pasadena": ("United States", "California"),
    "long island": ("United States", "New York"),
    "brooklyn": ("United States", "New York"),
    "queens": ("United States", "New York"),
    "manhattan": ("United States", "New York"),
    "bronx": ("United States", "New York"),
    "newark": ("United States", "New Jersey"),
    "jersey city": ("United States", "New Jersey"),
    "trenton": ("United States", "New Jersey"),
    "princeton": ("United States", "New Jersey"),
    "stamford": ("United States", "Connecticut"),
    "hartford": ("United States", "Connecticut"),
    "providence": ("United States", "Rhode Island"),
    "annandale": ("United States", "Virginia"),
    "arlington": ("United States", "Virginia"),
    "alexandria": ("United States", "Virginia"),
    "reston": ("United States", "Virginia"),
    "tysons": ("United States", "Virginia"),
    "fairfax": ("United States", "Virginia"),
    "bethesda": ("United States", "Maryland"),
    "rockville": ("United States", "Maryland"),
    "annapolis": ("United States", "Maryland"),
    "silver spring": ("United States", "Maryland"),
    "ann arbor": ("United States", "Michigan"),
    "grand rapids": ("United States", "Michigan"),
    "madison": ("United States", "Wisconsin"),
    "des moines": ("United States", "Iowa"),
    "boulder": ("United States", "Colorado"),
    "boise": ("United States", "Idaho"),
    "anchorage": ("United States", "Alaska"),
    "honolulu": ("United States", "Hawaii"),
    "research triangle": ("United States", "North Carolina"),
    "durham": ("United States", "North Carolina"),
    "chapel hill": ("United States", "North Carolina"),
    "winston-salem": ("United States", "North Carolina"),
    "asheville": ("United States", "North Carolina"),
    "savannah": ("United States", "Georgia"),
    "augusta": ("United States", "Georgia"),
    "tallahassee": ("United States", "Florida"),
    "gainesville": ("United States", "Florida"),
    "boca raton": ("United States", "Florida"),
    "naples fl": ("United States", "Florida"),
    "huntsville": ("United States", "Alabama"),
    "birmingham al": ("United States", "Alabama"),

    # ---------- Canada ----------
    "toronto": ("Canada", "Ontario"),
    "greater toronto": ("Canada", "Ontario"),
    "gta": ("Canada", "Ontario"),
    "ottawa": ("Canada", "Ontario"),
    "mississauga": ("Canada", "Ontario"),
    "brampton": ("Canada", "Ontario"),
    "markham": ("Canada", "Ontario"),
    "vaughan": ("Canada", "Ontario"),
    "kitchener": ("Canada", "Ontario"),
    "waterloo": ("Canada", "Ontario"),
    "windsor on": ("Canada", "Ontario"),
    "montreal": ("Canada", "Québec"),
    "greater montreal": ("Canada", "Québec"),
    "laval": ("Canada", "Québec"),
    "quebec city": ("Canada", "Québec"),
    "sherbrooke": ("Canada", "Québec"),
    "trois-rivieres": ("Canada", "Québec"),
    "gatineau": ("Canada", "Québec"),
    "vancouver": ("Canada", "British Columbia"),
    "burnaby": ("Canada", "British Columbia"),
    "richmond bc": ("Canada", "British Columbia"),
    "calgary": ("Canada", "Alberta"),
    "edmonton": ("Canada", "Alberta"),
    "winnipeg": ("Canada", "Manitoba"),
    "halifax": ("Canada", "Nova Scotia"),
    "saskatoon": ("Canada", "Saskatchewan"),
    "regina": ("Canada", "Saskatchewan"),
    "st. john's": ("Canada", "Newfoundland and Labrador"),
    "fredericton": ("Canada", "New Brunswick"),
    "moncton": ("Canada", "New Brunswick"),
    "charlottetown": ("Canada", "Prince Edward Island"),
    "whitehorse": ("Canada", "Yukon Territory"),

    # ---------- United Kingdom ----------
    "london": ("United Kingdom", ""),
    "greater london": ("United Kingdom", ""),
    "manchester": ("United Kingdom", ""),
    "birmingham": ("United Kingdom", ""),
    "glasgow": ("United Kingdom", ""),
    "liverpool": ("United Kingdom", ""),
    "edinburgh": ("United Kingdom", ""),
    "bristol": ("United Kingdom", ""),
    "leeds": ("United Kingdom", ""),
    "sheffield": ("United Kingdom", ""),
    "cardiff": ("United Kingdom", ""),
    "belfast": ("United Kingdom", ""),
    "newcastle": ("United Kingdom", ""),
    "nottingham": ("United Kingdom", ""),
    "southampton": ("United Kingdom", ""),
    "aberdeen": ("United Kingdom", ""),
    "oxford": ("United Kingdom", ""),
    "brighton": ("United Kingdom", ""),

    # ---------- France ----------
    "paris": ("France", ""),
    "ile-de-france": ("France", ""),
    "toulouse": ("France", ""),
    "lyon": ("France", ""),
    "marseille": ("France", ""),
    "bordeaux": ("France", ""),
    "lille": ("France", ""),
    "nantes": ("France", ""),
    "strasbourg": ("France", ""),
    "rennes": ("France", ""),
    "grenoble": ("France", ""),
    "montpellier": ("France", ""),
    "nice cote d'azur": ("France", ""),

    # ---------- Germany ----------
    "berlin": ("Germany", ""),
    "munich": ("Germany", ""),
    "munchen": ("Germany", ""),
    "hamburg": ("Germany", ""),
    "frankfurt": ("Germany", ""),
    "stuttgart": ("Germany", ""),
    "cologne": ("Germany", ""),
    "koln": ("Germany", ""),
    "dusseldorf": ("Germany", ""),
    "leipzig": ("Germany", ""),
    "dresden": ("Germany", ""),
    "bremen": ("Germany", ""),
    "hannover": ("Germany", ""),
    "nuremberg": ("Germany", ""),
    "nurnberg": ("Germany", ""),

    # ---------- Italy ----------
    "rome": ("Italy", ""),
    "roma": ("Italy", ""),
    "milan": ("Italy", ""),
    "milano": ("Italy", ""),
    "naples": ("Italy", ""),
    "napoli": ("Italy", ""),
    "turin": ("Italy", ""),
    "torino": ("Italy", ""),
    "florence": ("Italy", ""),
    "firenze": ("Italy", ""),
    "venice": ("Italy", ""),
    "venezia": ("Italy", ""),
    "bologna": ("Italy", ""),
    "genoa": ("Italy", ""),
    "genova": ("Italy", ""),

    # ---------- Spain ----------
    "madrid": ("Spain", ""),
    "barcelona": ("Spain", ""),
    "seville": ("Spain", ""),
    "sevilla": ("Spain", ""),
    "valencia": ("Spain", ""),
    "bilbao": ("Spain", ""),
    "zaragoza": ("Spain", ""),
    "malaga": ("Spain", ""),

    # ---------- Netherlands / Belgium / Switzerland / Austria ----------
    "amsterdam": ("Netherlands", ""),
    "rotterdam": ("Netherlands", ""),
    "the hague": ("Netherlands", ""),
    "utrecht": ("Netherlands", ""),
    "eindhoven": ("Netherlands", ""),
    "brussels": ("Belgium", ""),
    "antwerp": ("Belgium", ""),
    "ghent": ("Belgium", ""),
    "zurich": ("Switzerland", ""),
    "geneva": ("Switzerland", ""),
    "basel": ("Switzerland", ""),
    "bern": ("Switzerland", ""),
    "lausanne": ("Switzerland", ""),
    "vienna": ("Austria", ""),
    "graz": ("Austria", ""),
    "salzburg": ("Austria", ""),

    # ---------- Nordics ----------
    "stockholm": ("Sweden", ""),
    "gothenburg": ("Sweden", ""),
    "malmo": ("Sweden", ""),
    "oslo": ("Norway", ""),
    "bergen": ("Norway", ""),
    "copenhagen": ("Denmark", ""),
    "aarhus": ("Denmark", ""),
    "helsinki": ("Finland", ""),
    "tampere": ("Finland", ""),
    "reykjavik": ("Iceland", ""),

    # ---------- Ireland / Portugal / Greece ----------
    "dublin": ("Ireland", ""),
    "cork": ("Ireland", ""),
    "galway": ("Ireland", ""),
    "lisbon": ("Portugal", ""),
    "porto": ("Portugal", ""),
    "athens": ("Greece", ""),
    "thessaloniki": ("Greece", ""),

    # ---------- Eastern Europe ----------
    "warsaw": ("Poland", ""),
    "krakow": ("Poland", ""),
    "wroclaw": ("Poland", ""),
    "poznan": ("Poland", ""),
    "gdansk": ("Poland", ""),
    "prague": ("Czech Republic", ""),
    "praha": ("Czech Republic", ""),
    "brno": ("Czech Republic", ""),
    "budapest": ("Hungary", ""),
    "bucharest": ("Romania", ""),
    "cluj-napoca": ("Romania", ""),
    "sofia": ("Bulgaria", ""),
    "belgrade": ("Serbia", ""),
    "zagreb": ("Croatia", ""),
    "ljubljana": ("Slovenia", ""),
    "bratislava": ("Slovakia", ""),
    "tallinn": ("Estonia", ""),
    "riga": ("Latvia", ""),
    "vilnius": ("Lithuania", ""),
    "moscow": ("Russia", ""),
    "saint petersburg": ("Russia", ""),
    "novosibirsk": ("Russia", ""),
    "kyiv": ("Ukraine", ""),
    "kiev": ("Ukraine", ""),
    "lviv": ("Ukraine", ""),
    "minsk": ("Belarus", ""),

    # ---------- Middle East ----------
    "istanbul": ("Turkey", ""),
    "ankara": ("Turkey", ""),
    "izmir": ("Turkey", ""),
    "tel aviv": ("Israel", ""),
    "jerusalem": ("Israel", ""),
    "haifa": ("Israel", ""),
    "dubai": ("United Arab Emirates", ""),
    "abu dhabi": ("United Arab Emirates", ""),
    "sharjah": ("United Arab Emirates", ""),
    "doha": ("Qatar", ""),
    "riyadh": ("Saudi Arabia", ""),
    "jeddah": ("Saudi Arabia", ""),
    "mecca": ("Saudi Arabia", ""),
    "kuwait city": ("Kuwait", ""),
    "manama": ("Bahrain", ""),
    "muscat": ("Oman", ""),
    "amman": ("Jordan", ""),
    "beirut": ("Lebanon", ""),
    "tehran": ("Iran (Islamic Republic of)", ""),

    # ---------- Africa ----------
    "cairo": ("Egypt", ""),
    "alexandria eg": ("Egypt", ""),
    "casablanca": ("Morocco", ""),
    "rabat": ("Morocco", ""),
    "marrakech": ("Morocco", ""),
    "tunis": ("Tunisia", ""),
    "algiers": ("Algeria", ""),
    "lagos": ("Nigeria", ""),
    "abuja": ("Nigeria", ""),
    "nairobi": ("Kenya", ""),
    "johannesburg": ("South Africa", ""),
    "cape town": ("South Africa", ""),
    "pretoria": ("South Africa", ""),
    "durban": ("South Africa", ""),
    "addis ababa": ("Ethiopia", ""),
    "accra": ("Ghana", ""),
    "dakar": ("Senegal", ""),

    # ---------- Asia ----------
    "tokyo": ("Japan", ""),
    "osaka": ("Japan", ""),
    "kyoto": ("Japan", ""),
    "yokohama": ("Japan", ""),
    "nagoya": ("Japan", ""),
    "sapporo": ("Japan", ""),
    "fukuoka": ("Japan", ""),
    "seoul": ("South Korea", ""),
    "busan": ("South Korea", ""),
    "incheon": ("South Korea", ""),
    "daegu": ("South Korea", ""),
    "beijing": ("China", ""),
    "shenzhen": ("China", ""),
    "guangzhou": ("China", ""),
    "chengdu": ("China", ""),
    "hangzhou": ("China", ""),
    "nanjing": ("China", ""),
    "xian": ("China", ""),
    "wuhan": ("China", ""),
    "tianjin": ("China", ""),
    "taipei": ("Taiwan", ""),
    "kaohsiung": ("Taiwan", ""),
    "taichung": ("Taiwan", ""),
    "bangkok": ("Thailand", ""),
    "chiang mai": ("Thailand", ""),
    "kuala lumpur": ("Malaysia", ""),
    "penang": ("Malaysia", ""),
    "jakarta": ("Indonesia", ""),
    "surabaya": ("Indonesia", ""),
    "bali": ("Indonesia", ""),
    "manila": ("Philippines", ""),
    "cebu": ("Philippines", ""),
    "ho chi minh city": ("Vietnam", ""),
    "ho chi minh": ("Vietnam", ""),
    "hanoi": ("Vietnam", ""),
    "saigon": ("Vietnam", ""),
    "phnom penh": ("Cambodia", ""),
    "yangon": ("Myanmar", ""),
    "mumbai": ("India", ""),
    "bombay": ("India", ""),
    "delhi": ("India", ""),
    "new delhi": ("India", ""),
    "national capital region india": ("India", ""),
    "ncr india": ("India", ""),
    "bangalore": ("India", ""),
    "bengaluru": ("India", ""),
    "hyderabad": ("India", ""),
    "chennai": ("India", ""),
    "madras": ("India", ""),
    "kolkata": ("India", ""),
    "calcutta": ("India", ""),
    "pune": ("India", ""),
    "ahmedabad": ("India", ""),
    "gurgaon": ("India", ""),
    "gurugram": ("India", ""),
    "noida": ("India", ""),
    "karachi": ("Pakistan", ""),
    "lahore": ("Pakistan", ""),
    "islamabad": ("Pakistan", ""),
    "dhaka": ("Bangladesh", ""),
    "colombo": ("Sri Lanka", ""),
    "kathmandu": ("Nepal", ""),

    # ---------- Oceania ----------
    "sydney": ("Australia", ""),
    "melbourne": ("Australia", ""),
    "brisbane": ("Australia", ""),
    "perth": ("Australia", ""),
    "adelaide": ("Australia", ""),
    "canberra": ("Australia", ""),
    "gold coast": ("Australia", ""),
    "auckland": ("New Zealand", ""),
    "wellington": ("New Zealand", ""),
    "christchurch": ("New Zealand", ""),

    # ---------- Latin America ----------
    "mexico city": ("Mexico", ""),
    "guadalajara": ("Mexico", ""),
    "monterrey": ("Mexico", ""),
    "puebla": ("Mexico", ""),
    "tijuana": ("Mexico", ""),
    "queretaro": ("Mexico", ""),
    "sao paulo": ("Brazil", ""),
    "rio de janeiro": ("Brazil", ""),
    "brasilia": ("Brazil", ""),
    "belo horizonte": ("Brazil", ""),
    "porto alegre": ("Brazil", ""),
    "curitiba": ("Brazil", ""),
    "salvador": ("Brazil", ""),
    "recife": ("Brazil", ""),
    "fortaleza": ("Brazil", ""),
    "buenos aires": ("Argentina", ""),
    "cordoba": ("Argentina", ""),
    "santiago": ("Chile", ""),
    "valparaiso": ("Chile", ""),
    "lima": ("Peru", ""),
    "bogota": ("Colombia", ""),
    "medellin": ("Colombia", ""),
    "cali": ("Colombia", ""),
    "caracas": ("Venezuela", ""),
    "quito": ("Ecuador", ""),
    "guayaquil": ("Ecuador", ""),
    "la paz": ("Bolivia", ""),
    "asuncion": ("Paraguay", ""),
    "montevideo": ("Uruguay", ""),
    "panama city": ("Panama", ""),
    "san jose costa rica": ("Costa Rica", ""),
    "havana": ("Cuba", ""),
    "san salvador": ("El Salvador", ""),
    "tegucigalpa": ("Honduras", ""),
    "managua": ("Nicaragua", ""),
    "guatemala city": ("Guatemala", ""),
    "santo domingo": ("Dominican Republic", ""),
    "san juan pr": ("Puerto Rico", ""),
}

# ---------------------------------------------------------------------------
# Email ccTLD → Country
# ---------------------------------------------------------------------------
# Used when no Country / State / Location data is available, by inspecting the
# email-address top-level domain. Generic TLDs (.com, .org, .net, .edu, .gov,
# .io, .ai, .co, .me, .tv) are DELIBERATELY OMITTED because they do not imply
# a country. Multi-segment TLDs (.co.uk, .com.au) are checked before single
# segments via longest-match ordering.
EMAIL_TLD_TO_COUNTRY: Dict[str, str] = {
    # Europe
    ".uk": "United Kingdom", ".co.uk": "United Kingdom",
    ".ac.uk": "United Kingdom", ".org.uk": "United Kingdom",
    ".gov.uk": "United Kingdom",
    ".ie": "Ireland", ".de": "Germany", ".fr": "France", ".it": "Italy",
    ".es": "Spain", ".pt": "Portugal", ".nl": "Netherlands", ".be": "Belgium",
    ".lu": "Luxembourg", ".ch": "Switzerland", ".at": "Austria",
    ".se": "Sweden", ".no": "Norway", ".dk": "Denmark", ".fi": "Finland",
    ".is": "Iceland", ".pl": "Poland", ".cz": "Czech Republic",
    ".sk": "Slovakia", ".hu": "Hungary", ".ro": "Romania", ".bg": "Bulgaria",
    ".gr": "Greece", ".hr": "Croatia", ".si": "Slovenia", ".rs": "Serbia",
    ".ee": "Estonia", ".lv": "Latvia", ".lt": "Lithuania",
    ".ru": "Russia", ".ua": "Ukraine", ".by": "Belarus", ".md": "Moldova",
    # Middle East
    ".tr": "Turkey", ".il": "Israel", ".sa": "Saudi Arabia",
    ".ae": "United Arab Emirates", ".qa": "Qatar", ".kw": "Kuwait",
    ".bh": "Bahrain", ".om": "Oman", ".jo": "Jordan", ".lb": "Lebanon",
    ".ir": "Iran (Islamic Republic of)",
    # Africa
    ".eg": "Egypt", ".ma": "Morocco", ".tn": "Tunisia", ".dz": "Algeria",
    ".za": "South Africa", ".ng": "Nigeria", ".ke": "Kenya", ".gh": "Ghana",
    ".sn": "Senegal", ".et": "Ethiopia",
    # Asia
    ".jp": "Japan", ".kr": "South Korea", ".cn": "China", ".com.cn": "China",
    ".tw": "Taiwan", ".com.tw": "Taiwan", ".hk": "Hong Kong",
    ".com.hk": "Hong Kong", ".sg": "Singapore", ".com.sg": "Singapore",
    ".my": "Malaysia", ".com.my": "Malaysia", ".id": "Indonesia",
    ".co.id": "Indonesia", ".th": "Thailand", ".co.th": "Thailand",
    ".vn": "Vietnam", ".com.vn": "Vietnam", ".ph": "Philippines",
    ".com.ph": "Philippines", ".in": "India", ".co.in": "India",
    ".pk": "Pakistan", ".com.pk": "Pakistan", ".bd": "Bangladesh",
    ".com.bd": "Bangladesh", ".lk": "Sri Lanka", ".np": "Nepal",
    ".kh": "Cambodia", ".mm": "Myanmar",
    # Oceania
    ".au": "Australia", ".com.au": "Australia",
    ".nz": "New Zealand", ".co.nz": "New Zealand",
    # Americas
    ".us": "United States", ".ca": "Canada",
    ".mx": "Mexico", ".com.mx": "Mexico",
    ".br": "Brazil", ".com.br": "Brazil",
    ".ar": "Argentina", ".com.ar": "Argentina",
    ".cl": "Chile", ".pe": "Peru", ".ve": "Venezuela",
    ".uy": "Uruguay", ".py": "Paraguay", ".bo": "Bolivia",
    ".ec": "Ecuador", ".cr": "Costa Rica", ".pa": "Panama",
    ".do": "Dominican Republic", ".gt": "Guatemala", ".cu": "Cuba",
}

# ---------------------------------------------------------------------------
# Curated Company → Country HQ lookup (last-resort fallback)
# ---------------------------------------------------------------------------
# Conservative: only globally-unambiguous multinationals. Matching is done
# after lowercasing and stripping common corporate suffixes (Inc, Ltd, GmbH...)
# so 'Microsoft Corporation' → 'microsoft' → United States.
COMPANY_HQ: Dict[str, str] = {
    # OPAL-RT itself (Arnaud's employer)
    "opal-rt": "Canada", "opal rt": "Canada",
    "opal-rt technologies": "Canada",
    # US tech / cloud
    "microsoft": "United States", "apple": "United States",
    "google": "United States", "alphabet": "United States",
    "amazon": "United States", "aws": "United States", "meta": "United States",
    "facebook": "United States", "netflix": "United States",
    "adobe": "United States", "oracle": "United States",
    "salesforce": "United States", "ibm": "United States",
    "intel": "United States", "nvidia": "United States", "amd": "United States",
    "qualcomm": "United States", "cisco": "United States",
    "hp": "United States", "hewlett-packard": "United States",
    "hpe": "United States", "dell": "United States", "vmware": "United States",
    "tesla": "United States", "spacex": "United States",
    "openai": "United States", "anthropic": "United States",
    "uber": "United States", "stripe": "United States",
    "snowflake": "United States", "databricks": "United States",
    "palantir": "United States",
    # US industrial / aerospace / defense / energy
    "boeing": "United States", "lockheed martin": "United States",
    "northrop grumman": "United States", "raytheon": "United States",
    "rtx": "United States", "general dynamics": "United States",
    "general electric": "United States", "ge": "United States",
    "ge vernova": "United States", "honeywell": "United States",
    "caterpillar": "United States", "john deere": "United States",
    "deere": "United States", "3m": "United States", "dow": "United States",
    "dupont": "United States", "exxonmobil": "United States",
    "chevron": "United States", "conocophillips": "United States",
    # US auto
    "ford": "United States", "general motors": "United States",
    "gm": "United States", "rivian": "United States",
    "lucid motors": "United States",
    # Germany
    "siemens": "Germany", "siemens energy": "Germany", "bosch": "Germany",
    "sap": "Germany", "bmw": "Germany", "mercedes-benz": "Germany",
    "mercedes": "Germany", "daimler": "Germany", "volkswagen": "Germany",
    "vw": "Germany", "audi": "Germany", "porsche": "Germany",
    "continental": "Germany", "zf": "Germany", "zf friedrichshafen": "Germany",
    "thyssenkrupp": "Germany", "lufthansa": "Germany", "deutsche bank": "Germany",
    "deutsche telekom": "Germany", "infineon": "Germany",
    # France
    "airbus": "France", "thales": "France", "safran": "France",
    "dassault": "France", "renault": "France", "stellantis": "France",
    "totalenergies": "France", "total": "France", "edf": "France",
    "schneider electric": "France", "schneider-electric": "France",
    "alstom": "France", "valeo": "France", "michelin": "France",
    "loreal": "France", "l'oréal": "France", "engie": "France",
    "ovh": "France", "ovhcloud": "France",
    # Italy
    "ferrari": "Italy", "fiat": "Italy", "stellantis italy": "Italy",
    "leonardo": "Italy", "eni": "Italy", "enel": "Italy",
    "pirelli": "Italy", "iveco": "Italy",
    # Other Europe
    "abb": "Switzerland", "nestle": "Switzerland", "nestlé": "Switzerland",
    "roche": "Switzerland", "novartis": "Switzerland", "ubs": "Switzerland",
    "credit suisse": "Switzerland", "shell": "Netherlands",
    "asml": "Netherlands", "philips": "Netherlands", "akzo nobel": "Netherlands",
    "heineken": "Netherlands", "ing": "Netherlands",
    "ericsson": "Sweden", "volvo": "Sweden", "ikea": "Sweden",
    "spotify": "Sweden", "scania": "Sweden", "saab": "Sweden",
    "atlas copco": "Sweden", "nokia": "Finland", "kone": "Finland",
    "fortum": "Finland", "equinor": "Norway", "statoil": "Norway",
    "novo nordisk": "Denmark", "maersk": "Denmark", "vestas": "Denmark",
    "ørsted": "Denmark", "orsted": "Denmark",
    "bp": "United Kingdom", "british petroleum": "United Kingdom",
    "rolls-royce": "United Kingdom", "rolls royce": "United Kingdom",
    "bae systems": "United Kingdom", "arm": "United Kingdom",
    "barclays": "United Kingdom", "hsbc": "United Kingdom",
    "vodafone": "United Kingdom", "unilever": "United Kingdom",
    "diageo": "United Kingdom", "glaxosmithkline": "United Kingdom",
    "gsk": "United Kingdom", "astrazeneca": "United Kingdom",
    "iberdrola": "Spain", "telefonica": "Spain", "telefónica": "Spain",
    "santander": "Spain", "repsol": "Spain", "indra": "Spain",
    # Japan
    "toyota": "Japan", "honda": "Japan", "nissan": "Japan", "mazda": "Japan",
    "subaru": "Japan", "suzuki": "Japan", "mitsubishi": "Japan",
    "mitsubishi electric": "Japan", "mitsubishi heavy industries": "Japan",
    "hitachi": "Japan", "sony": "Japan", "panasonic": "Japan",
    "fujitsu": "Japan", "nec": "Japan", "denso": "Japan", "yamaha": "Japan",
    "canon": "Japan", "nikon": "Japan", "olympus": "Japan",
    "softbank": "Japan", "rakuten": "Japan",
    # South Korea
    "samsung": "South Korea", "samsung electronics": "South Korea",
    "lg": "South Korea", "lg electronics": "South Korea",
    "hyundai": "South Korea", "kia": "South Korea", "sk hynix": "South Korea",
    "posco": "South Korea", "doosan": "South Korea",
    # China
    "huawei": "China", "alibaba": "China", "tencent": "China",
    "baidu": "China", "byd": "China", "didi": "China", "xiaomi": "China",
    "lenovo": "China", "haier": "China", "sinopec": "China",
    "petrochina": "China", "cnpc": "China", "state grid": "China",
    # Taiwan
    "tsmc": "Taiwan", "foxconn": "Taiwan", "asus": "Taiwan", "acer": "Taiwan",
    "mediatek": "Taiwan",
    # India
    "tata": "India", "tata motors": "India", "tata consultancy services": "India",
    "tcs": "India", "infosys": "India", "wipro": "India",
    "reliance industries": "India", "reliance": "India",
    "mahindra": "India", "bharat heavy electricals": "India",
    "bhel": "India",
    # Australia / Canada / Saudi Arabia / Brazil
    "bhp": "Australia", "rio tinto": "Australia",
    "qantas": "Australia", "telstra": "Australia",
    "bombardier": "Canada", "cae": "Canada", "magna": "Canada",
    "magna international": "Canada", "hydro-québec": "Canada",
    "hydro quebec": "Canada", "shopify": "Canada", "blackberry": "Canada",
    "rbc": "Canada", "td": "Canada", "td bank": "Canada", "cn rail": "Canada",
    "saudi aramco": "Saudi Arabia", "aramco": "Saudi Arabia",
    "sabic": "Saudi Arabia",
    "embraer": "Brazil", "petrobras": "Brazil", "vale": "Brazil",
}

EMAIL_REGEX = re.compile(
    r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
)

# Fields shown (and editable) in the column-mapping UI
MAPPABLE_TARGETS: List[str] = [
    "First Name",
    "Last Name",
    "Job Title",
    "Company Name",
    "Email",
    "Business Phone",
    "LinkedIn",
    "Country",
    "State or Province",
    "Description",
    "Market Segment",
    "Main Application",
    "Industry Sector",
]
# Pseudo-target shown alongside the above; if mapped, its value is parsed
# into Country + State/Province (instead of being copied verbatim).
LOCATION_TARGET_LABEL = "Location (parsed → Country + State/Province)"

# Brand palette
BRAND = {
    "navy": "#002F6C",
    "navy_dark": "#001A3F",
    "accent": "#0099D8",
    "accent_dark": "#007BB0",
    "bg": "#F4F6F9",
    "card": "#FFFFFF",
    "text": "#1A1A1A",
    "text_muted": "#5A6473",
    "border": "#E1E6EE",
    "success_bg": "#E8F6EC",
    "success_border": "#2E7D32",
    "error_bg": "#FDECEC",
    "error_border": "#C62828",
}

# ===========================================================================
# CUSTOM CSS / BRANDING
# ===========================================================================

CUSTOM_CSS = f"""
<style>
    /* ------- Base ------- */
    .stApp {{
        background: {BRAND['bg']};
    }}
    .block-container {{
        padding-top: 0rem !important;
        padding-bottom: 3rem !important;
        max-width: 1200px;
    }}
    /* Hide Streamlit chrome we don't need */
    #MainMenu {{visibility: hidden;}}
    footer {{visibility: hidden;}}
    header[data-testid="stHeader"] {{
        background: transparent;
    }}

    /* ------- Typography ------- */
    html, body, [class*="css"] {{
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Inter, Roboto, Helvetica, Arial, sans-serif;
        color: {BRAND['text']};
    }}

    /* ------- Hero ------- */
    .hero {{
        position: relative;
        width: 100%;
        min-height: 220px;
        border-radius: 16px;
        overflow: hidden;
        margin-top: 1rem;
        margin-bottom: 1.5rem;
        background:
            linear-gradient(120deg, rgba(0,26,63,0.85) 0%, rgba(0,47,108,0.65) 50%, rgba(0,153,216,0.45) 100%),
            url("https://www.opal-rt.com/wp-content/uploads/2025/05/Hero-News-OPAL-RT.jpg") center/cover no-repeat,
            linear-gradient(120deg, {BRAND['navy_dark']} 0%, {BRAND['navy']} 50%, {BRAND['accent']} 100%);
        color: white;
        padding: 2.25rem 2rem;
        display: flex;
        flex-direction: column;
        justify-content: center;
        box-shadow: 0 10px 30px rgba(0, 47, 108, 0.18);
    }}
    .hero-eyebrow {{
        font-size: 0.75rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        opacity: 0.85;
        margin-bottom: 0.5rem;
        font-weight: 600;
    }}
    .hero-title {{
        font-size: 2.2rem;
        font-weight: 700;
        margin: 0;
        line-height: 1.15;
        letter-spacing: -0.01em;
    }}
    .hero-subtitle {{
        font-size: 1.05rem;
        margin-top: 0.6rem;
        opacity: 0.95;
        max-width: 720px;
        line-height: 1.5;
    }}
    @media (max-width: 640px) {{
        .hero {{ padding: 1.5rem 1.25rem; min-height: 180px; border-radius: 12px; }}
        .hero-title {{ font-size: 1.5rem; }}
        .hero-subtitle {{ font-size: 0.95rem; }}
    }}

    /* ------- Section cards ------- */
    .section-card {{
        background: {BRAND['card']};
        border: 1px solid {BRAND['border']};
        border-radius: 14px;
        padding: 1.5rem 1.5rem 1.25rem;
        margin-bottom: 1.25rem;
        box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
    }}
    .section-card h3 {{
        margin: 0 0 0.25rem 0;
        font-size: 1.15rem;
        color: {BRAND['navy']};
        font-weight: 700;
    }}
    .section-card .section-hint {{
        font-size: 0.85rem;
        color: {BRAND['text_muted']};
        margin-bottom: 1rem;
    }}
    .req-asterisk {{
        color: {BRAND['accent']};
        font-weight: 700;
        margin-left: 2px;
    }}

    /* ------- Streamlit widgets ------- */
    /* Buttons */
    .stButton > button {{
        background: linear-gradient(180deg, {BRAND['accent']} 0%, {BRAND['accent_dark']} 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.6rem 1.25rem;
        font-weight: 600;
        font-size: 0.95rem;
        box-shadow: 0 2px 6px rgba(0, 153, 216, 0.25);
        transition: transform 0.04s ease, box-shadow 0.15s ease, filter 0.15s ease;
        width: 100%;
    }}
    .stButton > button:hover {{
        filter: brightness(1.05);
        box-shadow: 0 4px 14px rgba(0, 153, 216, 0.35);
    }}
    .stButton > button:active {{
        transform: translateY(1px);
    }}
    /* Download button (secondary look – navy) */
    .stDownloadButton > button {{
        background: linear-gradient(180deg, {BRAND['navy']} 0%, {BRAND['navy_dark']} 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.7rem 1.25rem;
        font-weight: 600;
        font-size: 1rem;
        width: 100%;
        box-shadow: 0 2px 6px rgba(0, 47, 108, 0.25);
    }}
    .stDownloadButton > button:hover {{
        filter: brightness(1.1);
    }}

    /* File uploader */
    [data-testid="stFileUploader"] section {{
        border: 2px dashed {BRAND['accent']};
        background: rgba(0, 153, 216, 0.04);
        border-radius: 12px;
        padding: 1rem;
    }}
    [data-testid="stFileUploader"] section:hover {{
        background: rgba(0, 153, 216, 0.08);
    }}

    /* Inputs / selects */
    .stTextInput input, .stTextArea textarea {{
        border-radius: 8px !important;
        border-color: {BRAND['border']} !important;
    }}
    .stTextInput input:focus, .stTextArea textarea:focus {{
        border-color: {BRAND['accent']} !important;
        box-shadow: 0 0 0 1px {BRAND['accent']} !important;
    }}
    /* Selectbox styling – keep the native dropdown chrome but tone the border */
    div[data-baseweb="select"] > div {{
        border-radius: 8px !important;
        border-color: {BRAND['border']} !important;
    }}

    /* Tables */
    .stDataFrame {{
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid {BRAND['border']};
    }}

    /* Custom status banners */
    .success-banner {{
        background: {BRAND['success_bg']};
        border-left: 5px solid {BRAND['success_border']};
        color: #1B5E20;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        font-weight: 600;
        margin: 0.5rem 0 1rem 0;
    }}
    .error-banner {{
        background: {BRAND['error_bg']};
        border-left: 5px solid {BRAND['error_border']};
        color: #8B0000;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        font-weight: 600;
        margin: 0.5rem 0 1rem 0;
    }}
    .stats-row {{
        display: flex;
        gap: 1rem;
        flex-wrap: wrap;
        margin: 0.5rem 0 1.25rem 0;
    }}
    .stat-pill {{
        background: white;
        border: 1px solid {BRAND['border']};
        border-radius: 10px;
        padding: 0.65rem 1rem;
        min-width: 120px;
        flex: 1 1 120px;
    }}
    .stat-pill .stat-label {{
        font-size: 0.72rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        color: {BRAND['text_muted']};
        font-weight: 600;
    }}
    .stat-pill .stat-value {{
        font-size: 1.4rem;
        font-weight: 700;
        color: {BRAND['navy']};
        margin-top: 0.15rem;
    }}

    /* Mapping list */
    .mapping-grid {{
        display: grid;
        grid-template-columns: 1fr;
        gap: 0.4rem;
        font-size: 0.9rem;
    }}
    @media (min-width: 720px) {{
        .mapping-grid {{ grid-template-columns: 1fr 1fr; }}
    }}
    .mapping-row {{
        background: #F8FAFC;
        border: 1px solid {BRAND['border']};
        border-radius: 8px;
        padding: 0.5rem 0.75rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }}
    .mapping-arrow {{ color: {BRAND['accent']}; font-weight: 700; }}
    .mapping-source {{ color: {BRAND['text_muted']}; }}
    .mapping-target {{ color: {BRAND['navy']}; font-weight: 600; }}
    .mapping-missing {{ color: #B26A00; font-style: italic; }}

    /* Footer */
    .footer {{
        text-align: center;
        margin-top: 2.5rem;
        padding: 1.25rem;
        color: {BRAND['text_muted']};
        font-size: 0.85rem;
        border-top: 1px solid {BRAND['border']};
    }}
    .footer a {{
        color: {BRAND['accent']};
        text-decoration: none;
        font-weight: 600;
    }}
    .footer a:hover {{ text-decoration: underline; }}

    /* Make columns stack on small screens */
    @media (max-width: 640px) {{
        [data-testid="column"] {{
            width: 100% !important;
            flex: 1 1 100% !important;
        }}
    }}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ===========================================================================
# HELPER FUNCTIONS
# ===========================================================================

def fix_encoding(value) -> str:
    """Repair common mojibake (UTF-8 decoded as Latin-1) and strip junk chars."""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)

    s = value

    # Try the classic latin1→utf8 roundtrip if we see telltale "Ã" patterns
    if "Ã" in s or "Â" in s:
        try:
            fixed = s.encode("latin-1", errors="strict").decode("utf-8", errors="strict")
            # Only accept the fix if it removed the suspicious sequences
            if ("Ã" in s and "Ã" not in fixed) or ("Â" in s and "Â" not in fixed):
                s = fixed
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass

    # Hand-roll fallback for the most common Latin-letter mojibake patterns
    # that survive the roundtrip. Smart-quote / dash patterns are handled by
    # the encode/decode attempt above (we skip them here to keep the literals
    # plain ASCII-safe).
    mojibake_map = {
        "Ã©": "é", "Ã¨": "è", "Ãª": "ê", "Ã«": "ë",
        "Ã ": "à", "Ã¢": "â", "Ã¤": "ä", "Ã¡": "á", "Ã£": "ã",
        "Ã®": "î", "Ã¯": "ï", "Ã­": "í",
        "Ã´": "ô", "Ã¶": "ö", "Ã²": "ò", "Ã³": "ó",
        "Ã»": "û", "Ã¼": "ü", "Ã¹": "ù", "Ãº": "ú",
        "Ã±": "ñ", "Ã§": "ç", "Ã¿": "ÿ",
        "Ã‰": "É", "Ãˆ": "È", "ÃŠ": "Ê",
        "Ã€": "À", "Ã‚": "Â",
        "Ã\u201d": "Ô", "Ã™": "Ù", "Ã›": "Û",
        "Ã‡": "Ç", "Ã\u2018": "Ñ",
    }
    for bad, good in mojibake_map.items():
        if bad in s:
            s = s.replace(bad, good)

    # Strip replacement characters & zero-width junk
    s = s.replace("\ufffd", "")  # U+FFFD replacement
    s = s.replace("\u200b", "")  # zero-width space
    s = s.replace("\ufeff", "")  # BOM
    s = s.replace("\u00a0", " ")  # NBSP → space

    # Normalise unicode (combining accents → composed)
    s = unicodedata.normalize("NFC", s)
    return s


def _ascii_fold(s: str) -> str:
    """Strip accents/diacritics for substring matching.
    'São Paulo' → 'Sao Paulo', 'Montréal' → 'Montreal', 'Düsseldorf' → 'Dusseldorf'."""
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(c)
    )


def lookup_city_in_location(location: str) -> Tuple[str, str]:
    """Search a free-text location for any city in CITY_TO_GEO and return
    (country, state_or_province). Used as a final fallback for LinkedIn-style
    strings like 'Greater Chicago Area' or 'Greater Toulouse Metropolitan Area'
    where no country is explicitly named.

    Matches on word boundaries and prefers the longest city name (so 'New York'
    wins over 'York', 'Saint Petersburg' wins over 'Petersburg' if present)."""
    if not location:
        return "", ""
    haystack = _ascii_fold(location.lower())
    # Sort keys longest-first so multi-word cities win over single-word substrings
    for city in sorted(CITY_TO_GEO.keys(), key=len, reverse=True):
        if re.search(r"\b" + re.escape(city) + r"\b", haystack):
            return CITY_TO_GEO[city]
    return "", ""


def country_from_email(email: str) -> str:
    """Infer country from email domain's country-code TLD.
    'arnaud@opal-rt.com' → '' (generic .com, no signal)
    'jens@example.dk' → 'Denmark'
    'pierre@enterprise.co.uk' → 'United Kingdom'
    """
    if not email or "@" not in email:
        return ""
    domain = email.rsplit("@", 1)[1].lower().strip(" .,")
    if not domain:
        return ""
    # Longest-match first so '.co.uk' wins over '.uk'
    for tld in sorted(EMAIL_TLD_TO_COUNTRY.keys(), key=len, reverse=True):
        if domain.endswith(tld):
            return EMAIL_TLD_TO_COUNTRY[tld]
    return ""


_CORP_SUFFIX_PATTERN = re.compile(
    r"[,\s]*\b("
    r"inc|incorporated|llc|l\.l\.c\.?|ltd|limited|"
    r"corp|corporation|co|company|"
    r"plc|gmbh|ag|sa|s\.a\.?|sas|sarl|spa|s\.p\.a\.?|bv|b\.v\.?|nv|n\.v\.?|"
    r"oy|ab|as|kk|k\.k\.?|pty|pte|holdings|group|technologies|tech|"
    r"international|intl|systems|solutions|industries"
    r")\b\.?",
    re.IGNORECASE,
)


def _normalize_company(name: str) -> str:
    """Strip common corporate suffixes for fuzzy company-HQ lookup."""
    if not name:
        return ""
    s = name.lower().strip()
    # Repeatedly strip suffixes ("Microsoft Corp Inc" → "Microsoft Corp" → "Microsoft")
    prev = None
    while prev != s:
        prev = s
        s = _CORP_SUFFIX_PATTERN.sub("", s).strip(" ,.;-")
    s = re.sub(r"\s+", " ", s)
    return s


def country_from_company(company: str) -> str:
    """Best-effort lookup of company HQ → country.
    Tries the full suffix-stripped name first, then progressively drops
    trailing words (so 'Toyota Motor Corporation' → 'toyota motor' →
    'toyota' all get a chance to match). Returns '' if nothing matches."""
    norm = _normalize_company(company)
    if not norm:
        return ""
    parts = norm.split()
    while parts:
        candidate = " ".join(parts)
        if candidate in COMPANY_HQ:
            return COMPANY_HQ[candidate]
        folded = _ascii_fold(candidate)
        if folded != candidate and folded in COMPANY_HQ:
            return COMPANY_HQ[folded]
        parts.pop()  # drop the last word and try again
    return ""


def clean_text(value, lowercase: bool = False) -> str:
    """Trim, collapse whitespace, fix encoding."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = fix_encoding(str(value))
    s = re.sub(r"\s+", " ", s).strip()
    if lowercase:
        s = s.lower()
    return s


def normalize_header(name: str) -> str:
    """Normalize a column header for fuzzy matching."""
    if name is None:
        return ""
    s = fix_encoding(str(name)).lower().strip()
    # Drop non-alphanumerics for comparison
    return re.sub(r"[^a-z0-9]+", "", s)


def _column_data_quality(df: pd.DataFrame, col: str) -> int:
    """Return the number of non-empty values in df[col]. Used as a tiebreaker
    when multiple source columns match the same target (e.g., 'Account' and
    '#Account' — we pick whichever has more populated rows)."""
    if col not in df.columns:
        return 0
    series = df[col]
    # Pandas can return a DataFrame if there are duplicate column names —
    # collapse to the first such column in that case.
    if isinstance(series, pd.DataFrame):
        series = series.iloc[:, 0]
    s = series.astype(str).str.strip()
    return int(((s != "") & (s.str.lower() != "nan")).sum())


def detect_source_column(df: pd.DataFrame, target_field: str) -> Optional[str]:
    """Find the best matching source column for a Dynamics target field.
    When multiple source columns match the same target's aliases, return the
    one with the most non-empty values."""
    aliases = COLUMN_ALIASES.get(target_field, [])
    alias_keys = {normalize_header(a) for a in aliases if normalize_header(a)}
    if not alias_keys:
        return None

    # Keep ALL columns with their normalized forms (no dict-collision data loss)
    columns_with_norms: List[Tuple[str, str]] = [
        (normalize_header(c), c) for c in df.columns
    ]

    exact_candidates: List[str] = []
    substring_candidates: List[str] = []
    for norm_src, orig_src in columns_with_norms:
        if not norm_src:
            continue
        if norm_src in alias_keys:
            if orig_src not in exact_candidates:
                exact_candidates.append(orig_src)
        else:
            for key in alias_keys:
                if key in norm_src:
                    if orig_src not in substring_candidates:
                        substring_candidates.append(orig_src)
                    break

    pool = exact_candidates or substring_candidates
    if not pool:
        return None
    if len(pool) == 1:
        return pool[0]
    # Pick the candidate with the most populated rows
    return max(pool, key=lambda c: _column_data_quality(df, c))


def detect_location_column(df: pd.DataFrame) -> Optional[str]:
    """Locate a 'Location'-style column for parsing. Picks the richest one
    when multiple location-like columns exist."""
    alias_keys = {normalize_header(a) for a in LOCATION_ALIASES if normalize_header(a)}
    columns_with_norms = [(normalize_header(c), c) for c in df.columns]

    exact: List[str] = []
    contains: List[str] = []
    for norm_src, orig_src in columns_with_norms:
        if not norm_src:
            continue
        if norm_src in alias_keys:
            if orig_src not in exact:
                exact.append(orig_src)
        elif "location" in norm_src and orig_src not in contains:
            contains.append(orig_src)
    pool = exact or contains
    if not pool:
        return None
    if len(pool) == 1:
        return pool[0]
    return max(pool, key=lambda c: _column_data_quality(df, c))


def normalize_country(raw: str) -> str:
    """Return canonical country name (from COUNTRIES list) or '' if unrecognised."""
    if not raw:
        return ""
    s = clean_text(raw)
    if not s:
        return ""
    low = s.lower().strip(" .,")
    # Direct alias hit
    if low in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[low]
    # Case-insensitive direct match against canonical list
    for c in COUNTRIES:
        if c.lower() == low:
            return c
    # Try stripping common prefixes ("the ")
    if low.startswith("the "):
        return normalize_country(low[4:])
    return ""


def normalize_us_state(raw: str) -> str:
    """Return canonical US state name or '' if not a recognised US state."""
    if not raw:
        return ""
    s = clean_text(raw).strip(" .,")
    if not s:
        return ""
    low = s.lower()
    # Abbreviation (e.g. 'ca', 'n.y.')
    abbr_key = re.sub(r"[^a-z]", "", low)
    if abbr_key in US_STATE_ABBR:
        return US_STATE_ABBR[abbr_key]
    # Full name (case-insensitive)
    for st_name in US_STATES_FULL:
        if st_name.lower() == low:
            return st_name
    return ""


def normalize_ca_province(raw: str) -> str:
    """Return canonical Canadian province (with template spelling) or ''."""
    if not raw:
        return ""
    s = clean_text(raw).strip(" .,")
    if not s:
        return ""
    low = s.lower()
    abbr_key = re.sub(r"[^a-z]", "", low)
    if abbr_key in CA_PROVINCE_ABBR:
        return CA_PROVINCE_ABBR[abbr_key]
    for prov in CA_PROVINCES_FULL:
        if prov.lower() == low:
            return prov
    # Also accept 'Quebec' without accent
    if low == "quebec":
        return "Québec"
    return ""


def parse_location_string(loc: str) -> Tuple[str, str]:
    """Parse a free-text location into (country, state_or_province).
    Only US/Canada keep a state/province. Other countries return ('Country', '').
    Returns ('', '') if unable to determine confidently."""
    if not loc:
        return "", ""
    s = clean_text(loc)
    if not s:
        return "", ""

    # Normalise common alternate separators to commas so we can split uniformly.
    # Handles pipe, slash, semicolon, newline, and " - " separators.
    s_norm = re.sub(r"\s*[\|/;]\s*", ", ", s)
    s_norm = re.sub(r"\s*\n\s*", ", ", s_norm)
    s_norm = re.sub(r"\s+-\s+", ", ", s_norm)

    # Split on commas (and clean parts)
    parts = [p.strip() for p in s_norm.split(",") if p.strip()]

    if not parts:
        return "", ""

    # Try identifying a country anywhere in the parts (rightmost first)
    country = ""
    country_idx = -1
    for i in range(len(parts) - 1, -1, -1):
        c = normalize_country(parts[i])
        if c:
            country = c
            country_idx = i
            break

    state = ""

    if country == "United States":
        # Look for a US state in parts before the country position (or anywhere)
        search_range = parts[:country_idx] if country_idx >= 0 else parts
        for i in range(len(search_range) - 1, -1, -1):
            cand = normalize_us_state(search_range[i])
            if cand:
                state = cand
                break
    elif country == "Canada":
        search_range = parts[:country_idx] if country_idx >= 0 else parts
        for i in range(len(search_range) - 1, -1, -1):
            cand = normalize_ca_province(search_range[i])
            if cand:
                state = cand
                break

    # If no country found, try inferring from a US state or CA province
    if not country:
        for i in range(len(parts) - 1, -1, -1):
            cand = normalize_us_state(parts[i])
            if cand:
                country = "United States"
                state = cand
                break
            cand = normalize_ca_province(parts[i])
            if cand:
                country = "Canada"
                state = cand
                break

    # Last-resort substring scan for free-text locations like
    # "Greater New York City Area" or "San Francisco Bay Area, USA"
    if not country:
        low = s.lower()
        # Country aliases (longest first to avoid 'us' matching inside 'austin')
        for alias, canon in sorted(COUNTRY_ALIASES.items(), key=lambda kv: -len(kv[0])):
            if len(alias) < 4:
                continue  # avoid short tokens like 'us', 'uk' as substrings
            if re.search(r"\b" + re.escape(alias) + r"\b", low):
                country = canon
                break
        if not country:
            for canon in sorted(COUNTRIES, key=len, reverse=True):
                if len(canon) < 4:
                    continue
                if re.search(r"\b" + re.escape(canon.lower()) + r"\b", low):
                    country = canon
                    break

    # CITY lookup runs BEFORE the single-word state substring scan, because
    # multi-word city keys ('Washington DC', 'New York City', 'Tel Aviv') are
    # more specific than ambiguous single-word state names ('Washington').
    if not country:
        city_country, city_state = lookup_city_in_location(s)
        if city_country:
            country = city_country
            if city_state and city_country in ("United States", "Canada"):
                state = city_state

    # If still no country but the string mentions a known US state / CA province as a word
    if not country:
        for st_name in US_STATES_FULL:
            if re.search(r"\b" + re.escape(st_name.lower()) + r"\b", s.lower()):
                country = "United States"
                state = st_name
                break
    if not country:
        for prov in CA_PROVINCES_FULL:
            if re.search(r"\b" + re.escape(prov.lower()) + r"\b", s.lower()):
                country = "Canada"
                state = prov
                break

    return country, state


def validate_email(email: str) -> bool:
    if not email:
        return False
    return bool(EMAIL_REGEX.match(email))


def default_subject() -> str:
    """Default 'YYYYMMProspection' subject for the current month."""
    return datetime.now().strftime("%Y%m") + "Prospection"


def truncate_to(value: str, max_len: int) -> str:
    """Currently we DO NOT truncate – we surface a validation error instead.
    This helper exists in case future config wants to enable truncation."""
    return value if len(value) <= max_len else value[:max_len]


# ===========================================================================
# CORE PIPELINE
# ===========================================================================

def read_uploaded_file(uploaded) -> pd.DataFrame:
    """Read a CSV or Excel file uploaded via Streamlit into a pandas DataFrame."""
    name = uploaded.name.lower()
    raw = uploaded.read()
    bio = io.BytesIO(raw)

    if name.endswith((".xlsx", ".xlsm", ".xls")):
        # openpyxl is the supported engine
        engine = "openpyxl"
        df = pd.read_excel(bio, engine=engine, dtype=str)
    elif name.endswith(".csv"):
        # Try a couple of encodings (UTF-8 first, then cp1252 for legacy exports)
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                bio.seek(0)
                df = pd.read_csv(bio, dtype=str, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode the CSV file with common encodings.")
    else:
        raise ValueError("Unsupported file format. Please upload .csv or .xlsx.")

    return df


def strip_ghost_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove ghost columns: anything fully empty. Unnamed columns are KEPT
    if they hold data (some exports leave headers blank or write placeholders
    like 'Column9' for valid country/email cells)."""
    drop_cols = []
    for col in df.columns:
        is_empty = df[col].isna().all() or (df[col].astype(str).str.strip() == "").all()
        if is_empty:
            drop_cols.append(col)
            continue
        # If it's an utterly nameless column AND essentially empty header AND looks
        # like a pandas-generated placeholder, we already kept it above because it
        # has data. We deliberately do NOT drop populated 'Unnamed: N' / 'ColumnN'
        # columns — the row-scan fallback in process_dataframe will mine them for
        # country / state values.
    return df.drop(columns=drop_cols)


def build_column_mapping(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """Map Dynamics target fields → source column names found in df.
    When multiple source columns match the same target, the one with the most
    populated rows wins."""
    mapping: Dict[str, Optional[str]] = {}
    for target in COLUMN_ALIASES.keys():
        mapping[target] = detect_source_column(df, target)
    mapping["__location__"] = detect_location_column(df)
    return mapping


def process_dataframe(
    raw_df: pd.DataFrame,
    settings: Dict[str, str],
    mapping_override: Optional[Dict[str, Optional[str]]] = None,
) -> Tuple[pd.DataFrame, List[Dict], Dict[str, Optional[str]], Dict[str, int]]:
    """Main pipeline. Returns:
        - output_df: ready-to-export DataFrame in DYNAMICS_COLUMNS order
        - errors: list of {row, field, issue} dicts
        - mapping: column mapping actually used (auto + user overrides)
        - stats: counts of rows in / out / dropped

    `mapping_override`, when supplied, replaces individual entries in the auto-
    detected mapping. A value of None for a key in `mapping_override` means
    'do not use any source column for this target' (explicit un-map)."""

    df = strip_ghost_columns(raw_df.copy())

    # Normalize header strings (preserve original for display but trim)
    df.columns = [fix_encoding(str(c)).strip() for c in df.columns]

    auto_mapping = build_column_mapping(df)
    if mapping_override is not None:
        # Merge: override wins for any key present in override (incl. explicit None)
        mapping: Dict[str, Optional[str]] = {**auto_mapping, **mapping_override}
        # Validate that every mapped source column actually exists in df
        for k, v in list(mapping.items()):
            if v is not None and v not in df.columns:
                mapping[k] = None
    else:
        mapping = auto_mapping

    # Build the set of source columns we've already mapped to a target field.
    # The row-scan fallback for country/state uses this to avoid scanning the
    # email / phone / name / etc. columns for country values (which would risk
    # false positives), and instead only mines truly unmapped columns
    # (e.g. 'Column9', 'Column11' from sloppy exports).
    mapped_source_columns = set()
    for _tgt, _src in mapping.items():
        if _src:
            mapped_source_columns.add(_src)

    output_rows: List[Dict] = []
    errors: List[Dict] = []
    seen_emails: Dict[str, int] = {}

    stats = {
        "rows_in": len(df),
        "rows_out": 0,
        "rows_skipped_no_email": 0,
        "rows_duplicate_email": 0,
    }

    for idx, source_row in df.iterrows():
        # Source-spreadsheet row number for user-facing messages (header = row 1)
        source_row_num = int(idx) + 2

        def pull(target: str) -> str:
            src_col = mapping.get(target)
            if src_col is None or src_col not in df.columns:
                return ""
            return clean_text(source_row[src_col])

        # --- pull each field ---
        first_name = pull("First Name")
        last_name = pull("Last Name")
        company = pull("Company Name")
        job_title = pull("Job Title")
        email = pull("Email").lower()
        phone = pull("Business Phone")
        linkedin = pull("LinkedIn")
        description = pull("Description")
        country_raw = pull("Country")
        state_raw = pull("State or Province")
        location_raw = ""
        loc_col = mapping.get("__location__")
        if loc_col and loc_col in df.columns:
            location_raw = clean_text(source_row[loc_col])

        # --- skip rows with no email entirely (per spec addendum) ---
        if not email:
            stats["rows_skipped_no_email"] += 1
            continue

        # --- de-duplicate by email (case-insensitive) ---
        if email in seen_emails:
            stats["rows_duplicate_email"] += 1
            continue
        seen_emails[email] = source_row_num

        # --- resolve Country & State/Province ---
        country = normalize_country(country_raw) if country_raw else ""
        state = ""
        # Only treat State/Province as valid for US / Canada
        if state_raw:
            us_candidate = normalize_us_state(state_raw)
            ca_candidate = normalize_ca_province(state_raw)
            if us_candidate:
                state = us_candidate
                if not country:
                    country = "United States"
            elif ca_candidate:
                state = ca_candidate
                if not country:
                    country = "Canada"

        # If still missing, try the Location column
        if (not country or not state) and location_raw:
            loc_country, loc_state = parse_location_string(location_raw)
            if not country and loc_country:
                country = loc_country
            if not state and loc_state and country in ("United States", "Canada"):
                state = loc_state

        # LAST-RESORT FALLBACK 2 of 4: scan every unmapped column in this row
        # for a value that *exactly* matches a country / US state / CA province.
        # This rescues sloppy exports where country lives in unnamed columns
        # like 'Column9' or 'Unnamed: 10' with no recognisable header.
        if not country or not state:
            for col in df.columns:
                if country and state:
                    break
                if col in mapped_source_columns:
                    continue
                if col not in source_row.index:
                    continue
                val = clean_text(source_row[col])
                if not val or len(val) > 60:
                    continue
                if not country:
                    c_match = normalize_country(val)
                    if c_match:
                        country = c_match
                        continue
                if not state:
                    us_match = normalize_us_state(val)
                    if us_match:
                        state = us_match
                        if not country:
                            country = "United States"
                        continue
                    ca_match = normalize_ca_province(val)
                    if ca_match:
                        state = ca_match
                        if not country:
                            country = "Canada"
                        continue

        # LAST-RESORT FALLBACK 3 of 4: country-code TLD on the email address.
        # 'jens@example.dk' → Denmark; 'pierre@enterprise.co.uk' → United Kingdom.
        # Generic TLDs (.com / .org / .net / .io / .ai) yield no signal.
        if not country and email:
            tld_country = country_from_email(email)
            if tld_country:
                country = tld_country

        # LAST-RESORT FALLBACK 4 of 4: curated company → HQ country lookup.
        # Only multinationals where the HQ is unambiguous (Microsoft → US,
        # Airbus → France, Toyota → Japan, etc.)
        if not country and company:
            hq_country = country_from_company(company)
            if hq_country:
                country = hq_country

        # Country may have been pulled but isn't on the canonical list — final guard
        if country and country not in COUNTRIES:
            # If it's actually a US state or CA province name that landed in the
            # country slot, recover gracefully
            recovered_state = normalize_us_state(country) or normalize_ca_province(country)
            if normalize_us_state(country):
                country, state = "United States", normalize_us_state(country)
            elif normalize_ca_province(country):
                country, state = "Canada", normalize_ca_province(country)
            else:
                country = ""

        # State only applies to US/Canada
        if country not in ("United States", "Canada"):
            state = ""

        # --- Marketing/segment/sector: user override (global setting) wins;
        #     otherwise pull from source-file column if mapped, but only if
        #     the value matches a canonical dropdown option (no invention).
        src_market_segment = clean_text(pull("Market Segment"))
        if src_market_segment not in MARKET_SEGMENT_OPTIONS:
            src_market_segment = ""
        market_segment = (
            settings.get("market_segment", "") or src_market_segment or ""
        )

        # Main Application depends on Market Segment — validate against the
        # allowed list for the resolved segment.
        src_main_application = clean_text(pull("Main Application"))
        allowed_apps = MAIN_APPLICATION_BY_SEGMENT.get(market_segment, [""])
        if src_main_application not in allowed_apps:
            src_main_application = ""
        main_application = (
            settings.get("main_application", "") or src_main_application or ""
        )
        if main_application not in allowed_apps:
            main_application = ""

        src_industry_sector = clean_text(pull("Industry Sector"))
        if src_industry_sector not in INDUSTRY_SECTOR_OPTIONS:
            src_industry_sector = ""
        industry_sector = (
            settings.get("industry_sector", "") or src_industry_sector or ""
        )

        # --- Subject (mandatory, from global settings) ---
        subject = settings.get("subject", "").strip() or default_subject()

        out = {
            "(Do Not Modify) Lead": "",
            "(Do Not Modify) Row Checksum": "",
            "(Do Not Modify) Modified On": "",
            "Subject": subject,
            "First Name": first_name,
            "Last Name": last_name,
            "Job Title": job_title,
            "Company Name": company,
            "Email": email,
            "Business Phone": phone,
            "Country": country,
            "State or Province": state,
            "Description": description or settings.get("description", ""),
            "Lead Source": settings.get("lead_source", ""),
            "Rating": settings.get("rating", ""),
            "Source Campaign": settings.get("source_campaign", ""),
            "Market Segment": market_segment,
            "Main Application": main_application,
            "Industry Sector": industry_sector,
            "LinkedIn": linkedin,
            "Allow Marketing Communication": settings.get("allow_marketing", ""),
        }

        # ----- Row-level validation -----
        # Mandatory checks
        for f in MANDATORY_FIELDS:
            if not out.get(f):
                # For Country, include diagnostic context so the user can see
                # which source values we tried and why they didn't resolve.
                if f == "Country":
                    diag_bits = []
                    if country_raw:
                        diag_bits.append(f"raw country '{country_raw[:40]}' not recognised")
                    if location_raw:
                        diag_bits.append(f"location '{location_raw[:60]}' not parseable")
                    if not country_raw and not location_raw:
                        diag_bits.append("no Country or Location column detected for this row")
                    diag = f" — {'; '.join(diag_bits)}" if diag_bits else ""
                    errors.append({
                        "row": source_row_num,
                        "field": f,
                        "issue": f"Missing required field → Country{diag}",
                    })
                else:
                    errors.append({
                        "row": source_row_num,
                        "field": f,
                        "issue": f"Missing required field → {f}",
                    })

        # Email format
        if email and not validate_email(email):
            errors.append({
                "row": source_row_num,
                "field": "Email",
                "issue": f"Invalid email → {email}",
            })

        # Length validation (errors only — we do not silently truncate)
        for f, max_len in FIELD_MAX_LENGTHS.items():
            val = out.get(f, "")
            if val and len(val) > max_len:
                errors.append({
                    "row": source_row_num,
                    "field": f,
                    "issue": f"{f} exceeds {max_len} characters ({len(val)})",
                })

        output_rows.append(out)

    output_df = pd.DataFrame(output_rows, columns=DYNAMICS_COLUMNS)
    stats["rows_out"] = len(output_df)
    return output_df, errors, mapping, stats


def _safe_filename_component(name: str) -> str:
    """Strip characters illegal in filenames on Windows / macOS / Linux."""
    if not name:
        return ""
    s = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "", str(name))
    s = re.sub(r"\s+", " ", s).strip(" .")
    return s


def build_download_filename(
    output_df: pd.DataFrame, fallback_subject: str = ""
) -> str:
    """Compose the download filename per spec: 'YYMMDD - <Subject>.xlsx'.
    Uses the Subject actually written into the export (so the filename and the
    data stay in sync), falling back to the user's current setting or the
    default 'YYYYMMProspection' if the export is empty."""
    today = datetime.now().strftime("%y%m%d")
    subject = ""
    if len(output_df) > 0 and "Subject" in output_df.columns:
        first_val = output_df["Subject"].iloc[0]
        if first_val and not pd.isna(first_val):
            subject = str(first_val).strip()
    if not subject:
        subject = (fallback_subject or "").strip() or default_subject()
    subject = _safe_filename_component(subject) or default_subject()
    return f"{today} - {subject}.xlsx"


TEMPLATE_FILENAME = "ImportLeadTemplate.xlsm"


def _template_path() -> str:
    """Path to the bundled Dynamics template, expected next to this script."""
    import os
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_FILENAME)


def template_available() -> bool:
    import os
    return os.path.isfile(_template_path())


def df_to_dynamics_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Serialize the export by INJECTING rows into the official Dynamics
    template (ImportLeadTemplate.xlsm) rather than building a new workbook.

    Dynamics 365's 'Import from Excel' validates internal metadata that only
    exists in files derived from its own template:
      - a signed entity-mapping string in hiddenSheet!A1
      - the hidden lookup sheets (hiddenMarketSegments, etc.)
      - ~249 defined names used by dependent dropdowns
      - the 'Table1' Excel table on the Lead sheet whose range marks the data
    A from-scratch workbook is rejected with error 0x800608c3 ('Invalid Format
    in Import File'). By loading the real template, clearing its data rows,
    writing ours, and updating Table1's range, all of that metadata survives
    and Dynamics accepts the upload.

    Layout facts (verified against the template):
      - Row 1 of 'Lead' is empty; headers are in ROW 2; data starts ROW 3.
      - Columns A..U = the 21 DYNAMICS_COLUMNS in order.
    Saved via openpyxl without keep_vba → output is macro-free .xlsx content.
    """
    import warnings as _warnings
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    with _warnings.catch_warnings():
        # The template carries x14 conditional-formatting / data-validation
        # extensions openpyxl can't rewrite; they're for human data entry and
        # not required by the Dynamics import validator.
        _warnings.simplefilter("ignore")
        wb = load_workbook(_template_path())

    ws = wb["Lead"]
    n_cols = len(DYNAMICS_COLUMNS)

    # --- Sanity-check the header row so we never write misaligned data ---
    header_row = 2
    template_headers = [
        ws.cell(row=header_row, column=c).value for c in range(1, n_cols + 1)
    ]
    if template_headers != DYNAMICS_COLUMNS:
        raise ValueError(
            "Template header mismatch — the bundled ImportLeadTemplate.xlsm "
            f"headers in row {header_row} do not match the expected Dynamics "
            "columns. Re-download the template from Dynamics and replace the "
            "bundled copy."
        )

    data_start = header_row + 1  # row 3

    # --- Clear any leftover data rows from the template ---
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).value = None

    # --- Write our rows ---
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start + i
        for j, header in enumerate(DYNAMICS_COLUMNS, start=1):
            val = row.get(header, "")
            if pd.isna(val):
                val = ""
            # Strings only: prevents Excel from mangling phone numbers,
            # checksums, leading zeros, etc.
            ws.cell(row=excel_row, column=j).value = (
                str(val) if str(val) != "" else None
            )

    # --- Update the Excel table range so Dynamics sees exactly our rows ---
    last_col_letter = get_column_letter(n_cols)
    # A table must span header + at least one data row
    last_data_row = max(data_start, header_row + len(df))
    if len(df) > 0:
        last_data_row = header_row + len(df)
    if "Table1" in ws.tables:
        ws.tables["Table1"].ref = f"A{header_row}:{last_col_letter}{last_data_row}"

    buf = io.BytesIO()
    with _warnings.catch_warnings():
        _warnings.simplefilter("ignore")
        wb.save(buf)
    return buf.getvalue()


def df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """FALLBACK generic XLSX builder (used only if the bundled template is
    missing). NOTE: Dynamics 365 'Import from Excel' will REJECT this output
    (error 0x800608c3) because it lacks the template's hidden metadata. It
    remains useful as a human-readable export.

    - Sheet name 'Lead'
    - Bold OPAL-RT-navy header row in row 1
    - Frozen top row + auto-filter
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = list(df.columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead"

    # ----- Styled header row -----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="002F6C")  # OPAL-RT navy
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ----- Data rows -----
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, header in enumerate(headers, start=1):
            val = row[header]
            if pd.isna(val):
                val = ""
            # Convert everything to string to avoid accidental type coercion
            # by Excel (e.g. phone numbers losing leading zeros)
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")

    # ----- Polish: freeze header, auto-filter, column widths -----
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(headers))
    last_row = max(len(df) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    ws.row_dimensions[1].height = 22

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        if len(df) > 0:
            try:
                data_max = df[header].astype(str).str.len().max()
                if pd.notna(data_max):
                    max_len = max(max_len, int(data_max))
            except Exception:
                pass
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 60)

    # ----- Serialize -----
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_xlsx_bytes(df: pd.DataFrame) -> Tuple[bytes, bool]:
    """Produce the download bytes. Returns (bytes, used_template).
    Prefers template injection (Dynamics-importable); falls back to the
    generic builder if the bundled template is missing or unreadable."""
    if template_available():
        try:
            return df_to_dynamics_xlsx_bytes(df), True
        except Exception:
            pass
    return df_to_xlsx_bytes(df), False


# ===========================================================================
# UI
# ===========================================================================

def render_hero() -> None:
    st.markdown(
        """
        <div class="hero">
            <div class="hero-eyebrow">OPAL-RT • Internal Tool</div>
            <h1 class="hero-title">Opal RT Spreadsheet Cleaner</h1>
            <p class="hero-subtitle">
                Prepare CRM-ready lead imports for Microsoft Dynamics.
                Auto-detect columns, normalize data, validate every row,
                and export a clean import-ready file in seconds.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_footer() -> None:
    st.markdown(
        """
        <div class="footer">
            Built by <strong>Arnaud Joakim</strong> ·
            <a href="mailto:arnaud.joakim@opal-rt.com">arnaud.joakim@opal-rt.com</a>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_global_settings() -> Dict[str, str]:
    """Render the 'Global Import Settings' section and return the chosen values."""
    st.markdown(
        '<div class="section-card">'
        '<h3>① Global Import Settings</h3>'
        '<div class="section-hint">These values are applied to <strong>every row</strong> of the export. '
        'Fields marked with <span class="req-asterisk">*</span> are required by Dynamics.</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        subject = st.text_input(
            "Subject *",
            value=default_subject(),
            max_chars=300,
            help="Default format: YYYYMMProspection (e.g. 202605Prospection)",
        )
        lead_source = st.selectbox("Lead Source", options=LEAD_SOURCE_OPTIONS, index=0)
        rating = st.selectbox("Rating", options=RATING_OPTIONS, index=0)
        allow_marketing = st.selectbox(
            "Allow Marketing Communication",
            options=ALLOW_MARKETING_OPTIONS,
            index=0,
        )
        source_campaign = st.text_input(
            "Source Campaign",
            value="",
            help="Free-text campaign identifier (optional)",
        )

    with col2:
        market_segment = st.selectbox(
            "Market Segment",
            options=MARKET_SEGMENT_OPTIONS,
            index=0,
            key="market_segment_select",
        )
        # Main Application options are dependent on Market Segment selection
        main_app_options = MAIN_APPLICATION_BY_SEGMENT.get(market_segment, [""])
        main_application = st.selectbox(
            "Main Application",
            options=main_app_options,
            index=0,
            key="main_application_select",
            help="Options change based on Market Segment selection",
        )
        industry_sector = st.selectbox(
            "Industry Sector",
            options=INDUSTRY_SECTOR_OPTIONS,
            index=0,
        )
        description = st.text_area(
            "Description",
            value="",
            height=98,
            help="Optional default description applied when source row has none",
        )

    st.markdown("</div>", unsafe_allow_html=True)

    return {
        "subject": subject.strip(),
        "lead_source": lead_source,
        "rating": rating,
        "allow_marketing": allow_marketing,
        "source_campaign": source_campaign.strip(),
        "market_segment": market_segment,
        "main_application": main_application,
        "industry_sector": industry_sector,
        "description": description.strip(),
    }


MAPPING_KEY_PREFIX = "map_select_"
LOCATION_TARGET_KEY = "__location__"


def _mapping_key(target: str) -> str:
    """Stable session_state key for a target field's selectbox."""
    return MAPPING_KEY_PREFIX + re.sub(r"[^A-Za-z0-9]+", "_", target).lower()


def initialise_mapping_state(df: pd.DataFrame, auto_mapping: Dict[str, Optional[str]]) -> None:
    """Seed session_state with the auto-detected mapping. Called once per
    uploaded file. Existing user overrides for *other* files are cleared by
    main() when a new file is detected."""
    options = ["(none)"] + list(df.columns)
    for target in MAPPABLE_TARGETS:
        key = _mapping_key(target)
        if key not in st.session_state:
            auto_val = auto_mapping.get(target)
            st.session_state[key] = auto_val if auto_val in options else "(none)"
    loc_key = _mapping_key(LOCATION_TARGET_KEY)
    if loc_key not in st.session_state:
        loc_auto = auto_mapping.get(LOCATION_TARGET_KEY)
        st.session_state[loc_key] = loc_auto if loc_auto in options else "(none)"


def render_mapping_editor(
    df: pd.DataFrame, auto_mapping: Dict[str, Optional[str]]
) -> None:
    """Editable column mapping with a Save button. Each Dynamics target field
    gets its own dropdown of source columns. Defaults come from auto-detection;
    the user can change any of them. Changes are live (no Apply needed); Save
    is a confirmation gesture."""
    options = ["(none)"] + list(df.columns)

    st.markdown(
        '<div class="section-card">'
        '<h3>③ Column mapping</h3>'
        '<div class="section-hint">Each Dynamics field is mapped to the best-'
        'matching source column. Adjust any mapping with the dropdowns below '
        "and click <strong>Save mapping</strong> to confirm. Choose "
        '<em>(none)</em> to leave a field unmapped. The <em>Location</em> '
        'slot, when set, is parsed into Country + State/Province.</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    half = (len(MAPPABLE_TARGETS) + 1) // 2
    for i, target in enumerate(MAPPABLE_TARGETS):
        target_with_star = (
            target + " *"
            if target in MANDATORY_FIELDS
            else target
        )
        container = col1 if i < half else col2
        with container:
            current = st.session_state.get(_mapping_key(target), "(none)")
            if current not in options:
                current = "(none)"
            st.selectbox(
                target_with_star,
                options=options,
                index=options.index(current),
                key=_mapping_key(target),
                help=(
                    f"Source column to use for '{target}'. "
                    f"Auto-detected: {auto_mapping.get(target) or '— none —'}"
                ),
            )

    # Location pseudo-target on its own line (full width)
    loc_current = st.session_state.get(_mapping_key(LOCATION_TARGET_KEY), "(none)")
    if loc_current not in options:
        loc_current = "(none)"
    st.selectbox(
        "Location (parsed → Country + State/Province)",
        options=options,
        index=options.index(loc_current),
        key=_mapping_key(LOCATION_TARGET_KEY),
        help=(
            "Free-text location column to parse into Country + State/Province. "
            f"Auto-detected: {auto_mapping.get(LOCATION_TARGET_KEY) or '— none —'}"
        ),
    )

    save_col, reset_col, _spacer = st.columns([1, 1, 2])
    with save_col:
        if st.button("💾 Save mapping", key="btn_save_mapping"):
            st.session_state["mapping_saved"] = True
    with reset_col:
        if st.button("↺ Reset to auto", key="btn_reset_mapping"):
            for t in MAPPABLE_TARGETS:
                st.session_state[_mapping_key(t)] = (
                    auto_mapping.get(t)
                    if auto_mapping.get(t) in options
                    else "(none)"
                )
            st.session_state[_mapping_key(LOCATION_TARGET_KEY)] = (
                auto_mapping.get(LOCATION_TARGET_KEY)
                if auto_mapping.get(LOCATION_TARGET_KEY) in options
                else "(none)"
            )
            st.session_state["mapping_saved"] = False
            st.rerun()

    if st.session_state.get("mapping_saved"):
        st.markdown(
            '<div class="success-banner" style="margin-top: 0.75rem;">'
            '✓ Mapping saved — click <strong>Process file</strong> below to apply it.'
            '</div>',
            unsafe_allow_html=True,
        )

    st.markdown("</div>", unsafe_allow_html=True)


def get_user_mapping() -> Dict[str, Optional[str]]:
    """Read the user's column-mapping selections from session_state."""
    mapping: Dict[str, Optional[str]] = {}
    for target in MAPPABLE_TARGETS:
        val = st.session_state.get(_mapping_key(target))
        mapping[target] = None if not val or val == "(none)" else val
    loc_val = st.session_state.get(_mapping_key(LOCATION_TARGET_KEY))
    mapping[LOCATION_TARGET_KEY] = None if not loc_val or loc_val == "(none)" else loc_val
    return mapping


def render_stats(stats: Dict[str, int], n_errors: int) -> None:
    st.markdown(
        f"""
        <div class="stats-row">
            <div class="stat-pill">
                <div class="stat-label">Rows in source</div>
                <div class="stat-value">{stats['rows_in']}</div>
            </div>
            <div class="stat-pill">
                <div class="stat-label">Rows exported</div>
                <div class="stat-value">{stats['rows_out']}</div>
            </div>
            <div class="stat-pill">
                <div class="stat-label">Skipped (no email)</div>
                <div class="stat-value">{stats['rows_skipped_no_email']}</div>
            </div>
            <div class="stat-pill">
                <div class="stat-label">Duplicates removed</div>
                <div class="stat-value">{stats['rows_duplicate_email']}</div>
            </div>
            <div class="stat-pill">
                <div class="stat-label">Validation issues</div>
                <div class="stat-value">{n_errors}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_errors(errors: List[Dict]) -> None:
    """Show row-level errors in a tidy red panel."""
    if not errors:
        return
    st.markdown(
        f'<div class="error-banner">⚠️ {len(errors)} validation issue(s) detected. '
        f'Row numbers refer to the <strong>source file</strong> '
        f'(row 1 = headers, row 2 = first data row). '
        f'You can still export — Dynamics may reject affected rows.</div>',
        unsafe_allow_html=True,
    )
    err_df = pd.DataFrame(errors).sort_values(["row", "field"]).reset_index(drop=True)
    err_df.columns = ["Source row", "Field", "Issue"]
    st.dataframe(err_df, use_container_width=True, hide_index=True)


# ===========================================================================
# MAIN APP
# ===========================================================================

def main() -> None:
    render_hero()

    # -------- Global settings panel --------
    settings = render_global_settings()

    # -------- File upload panel --------
    st.markdown(
        '<div class="section-card">'
        '<h3>② Upload Source File</h3>'
        '<div class="section-hint">CSV or Excel (.xlsx). The app will auto-detect '
        'and map columns to the Dynamics template — you can adjust the mapping '
        'before processing.</div>',
        unsafe_allow_html=True,
    )
    uploaded = st.file_uploader(
        "Upload CSV or Excel File",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=False,
        label_visibility="collapsed",
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if not uploaded:
        st.info(
            "📥 Upload a lead spreadsheet above to begin. "
            "Supported formats: **.csv** and **.xlsx**.",
            icon="ℹ️",
        )
        render_footer()
        return

    # -------- Read file + auto-detect mapping (once per upload) --------
    is_new_file = (
        st.session_state.get("uploaded_file_name") != uploaded.name
        or st.session_state.get("uploaded_file_size") != uploaded.size
    )
    if is_new_file:
        # Reset any prior mapping state so the new file's auto-detection takes effect
        for k in list(st.session_state.keys()):
            if k.startswith(MAPPING_KEY_PREFIX):
                del st.session_state[k]
        st.session_state["mapping_saved"] = False
        st.session_state.pop("processed", None)
        try:
            with st.spinner("Reading file…"):
                raw_df = read_uploaded_file(uploaded)
            cleaned = strip_ghost_columns(raw_df.copy())
            cleaned.columns = [fix_encoding(str(c)).strip() for c in cleaned.columns]
            st.session_state["cleaned_df"] = cleaned
            st.session_state["auto_mapping"] = build_column_mapping(cleaned)
            st.session_state["uploaded_file_name"] = uploaded.name
            st.session_state["uploaded_file_size"] = uploaded.size
        except Exception as e:  # noqa: BLE001
            st.markdown(
                f'<div class="error-banner">❌ Failed to read file: {e}</div>',
                unsafe_allow_html=True,
            )
            render_footer()
            return

    cleaned_df = st.session_state["cleaned_df"]
    auto_mapping = st.session_state["auto_mapping"]
    initialise_mapping_state(cleaned_df, auto_mapping)

    # -------- Editable column mapping --------
    render_mapping_editor(cleaned_df, auto_mapping)

    # -------- Process --------
    st.markdown(
        '<div class="section-card">'
        '<h3>④ Normalize & Validate</h3>'
        '<div class="section-hint">Click below to process the uploaded file with '
        'the current column mapping. The app will clean encoding, normalize '
        'formatting, parse locations, infer country (from email TLD or company '
        'HQ if needed), remove duplicate emails, and validate every row.</div>',
        unsafe_allow_html=True,
    )
    process_clicked = st.button("🚀 Process file", type="primary")
    st.markdown("</div>", unsafe_allow_html=True)

    if process_clicked:
        try:
            with st.spinner("Normalizing and validating data…"):
                user_mapping = get_user_mapping()
                output_df, errors, mapping, stats = process_dataframe(
                    cleaned_df, settings, mapping_override=user_mapping
                )
            st.session_state["processed"] = {
                "output_df": output_df,
                "errors": errors,
                "mapping": mapping,
                "stats": stats,
                "source_name": uploaded.name,
            }
        except Exception as e:  # noqa: BLE001
            st.markdown(
                f'<div class="error-banner">❌ Failed to process file: {e}</div>',
                unsafe_allow_html=True,
            )
            render_footer()
            return

    # -------- Results --------
    state = st.session_state.get("processed")
    if not state:
        render_footer()
        return

    output_df = state["output_df"]
    errors = state["errors"]
    stats = state["stats"]

    st.markdown(
        '<div class="section-card">'
        '<h3>⑤ Results</h3>',
        unsafe_allow_html=True,
    )

    render_stats(stats, len(errors))

    if not errors:
        st.markdown(
            '<div class="success-banner">✅ File successfully normalized and ready for Dynamics import.</div>',
            unsafe_allow_html=True,
        )
    else:
        render_errors(errors)

    # Preview
    st.markdown("**Preview (first 50 rows of the export):**")
    st.dataframe(output_df.head(50), use_container_width=True, hide_index=True)

    # Download
    xlsx_bytes, used_template = export_xlsx_bytes(output_df)
    download_filename = build_download_filename(
        output_df, fallback_subject=settings.get("subject", "")
    )
    if not used_template:
        st.markdown(
            '<div class="error-banner">⚠️ The bundled Dynamics template '
            f'({TEMPLATE_FILENAME}) was not found next to the app, so a generic '
            'XLSX was generated instead. Dynamics 365 will reject it with '
            '"Invalid Format in Import File" (0x800608c3). Add the template '
            'file to the app folder and redeploy to fix this.</div>',
            unsafe_allow_html=True,
        )
    st.download_button(
        label=f"⬇️ Download {download_filename}",
        data=xlsx_bytes,
        file_name=download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("</div>", unsafe_allow_html=True)
    render_footer()


if __name__ == "__main__":
    main()
